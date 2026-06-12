import requests, json, base64, sys

API = open('/app/frontend/.env').read().split('REACT_APP_BACKEND_URL=')[1].strip().split('\n')[0] + '/api'
T = requests.post(f'{API}/auth/login', json={'username':'teacher1','password':'khairat1'}).json()['token']
H = {'Authorization': f'Bearer {T}'}

# cleanup old test gradebooks
for g in requests.get(f'{API}/gradebooks', headers=H).json():
    requests.delete(f'{API}/gradebooks/{g["id"]}', headers=H)

# 1. create gradebook with full official names
names = [
    'البراء بن سعود بن محمد هلال الجابري',
    'الحارث بن سيف بن راشد سليم الحارثي',
    'المختار بن أحمد بن ناصر محمد الحبسي',
    'الوليد بن خالد بن عبدالله سالم المنذري',
    'وليد بن محمد بن عبيد سلوم الحبسي',
]
gb = requests.post(f'{API}/gradebooks', json={'grade':'الخامس','section':'3','students':names}, headers=H).json()
print('1. created gradebook:', gb['grade'], gb['section'], 'students:', len(gb['students']))

# 2. create quiz + question + simulate submissions with PARTIAL names
quiz = requests.post(f'{API}/quizzes', json={'title':'اختبار الوحدة الأولى','settings':{'secret_code':'GBTEST'}}, headers=H).json()
qid = quiz['id']
requests.post(f'{API}/quizzes/{qid}/questions', json={'type':'mcq','text':'سؤال؟','options':['أ','ب','ج','د'],'correct_answer':'أ','points':10,'time_limit':30}, headers=H)
requests.post(f'{API}/quizzes/{qid}/activate', headers=H)
requests.post(f'{API}/quizzes/{qid}/start', headers=H)

partials = [
    ('البراء الجابري', 'أ'),          # partial: first + last
    ('الحارث بن سيف', 'ب'),           # partial: first half
    ('مختار الحبسي', 'أ'),            # missing ال + last
    ('الوليد المنذري', 'أ'),          # partial
    ('طالب غريب تماما', 'ج'),         # no match expected
]
sub_ids = {}
for name, ans in partials:
    j = requests.post(f'{API}/quiz/join/GBTEST').json() if False else None
    r = requests.post(f'{API}/quiz/{qid}/join', json={'student_name':name,'grade':'الخامس','section':'3'}).json()
    sid = r['submission_id']
    sub_ids[name] = sid
    qs = requests.get(f'{API}/quiz/{qid}/questions/{sid}').json()
    q1 = qs['questions'][0]
    requests.post(f'{API}/quiz/{qid}/submit/{sid}', json={'answers':[{'question_id':q1['id'],'answer_text':ans}]})

# 3. smart match
m = requests.post(f'{API}/gradebooks/{gb["id"]}/match-quiz', json={'quiz_id':qid}, headers=H).json()
print('\n2. MATCHING RESULTS:')
st_map = {s['id']: s['name'] for s in m['students']}
ok = 0
for p in m['proposals']:
    matched = st_map.get(p['matched_student_id'], '✗ غير مطابق')
    print(f"   '{p['student_name']}' -> '{matched}' (ثقة {p['confidence']})")
    if p['student_name'] == 'طالب غريب تماما':
        ok += (p['matched_student_id'] is None)
    else:
        ok += (p['matched_student_id'] is not None)
print('   matching correct:', ok, '/ 5')

# 4. apply to semester 1, column q1
mappings = [{'submission_id': p['id'], 'student_id': p['matched_student_id']} for p in m['proposals'] if p['matched_student_id']]
a = requests.post(f'{API}/gradebooks/{gb["id"]}/apply-quiz', json={'quiz_id':qid,'semester':'1','column':'q1','mappings':mappings}, headers=H).json()
print('\n3. applied:', a)

gb2 = requests.get(f'{API}/gradebooks/{gb["id"]}', headers=H).json()
for s in gb2['students']:
    sc = gb2['scores']['1'].get(s['id'], {})
    if sc: print('   ', s['name'][:25], '→ q1 =', sc.get('q1'))

# 5. export excel
r = requests.get(f'{API}/gradebooks/{gb["id"]}/export', headers=H)
open('/tmp/exported_gb.xlsx','wb').write(r.content)
print('\n4. export size:', len(r.content), 'bytes, content-type ok:', 'spreadsheet' in r.headers['content-type'])

# 6. verify exported structure with openpyxl
import openpyxl
wb = openpyxl.load_workbook('/tmp/exported_gb.xlsx')
print('   sheets:', wb.sheetnames)
ws = wb['الفصل الأول']
print('   A1:', ws['A1'].value[:50])
print('   B6:', ws['B6'].value, '| F6 (q1):', ws['F6'].value, '| O6 formula:', ws['O6'].value)
wsa = wb['سجل سنوي']
print('   annual C8:', wsa['C8'].value[:70])

# 7. import round-trip into a NEW gradebook (delete then import)
requests.delete(f'{API}/gradebooks/{gb["id"]}', headers=H)
b64 = base64.b64encode(open('/tmp/exported_gb.xlsx','rb').read()).decode()
imp = requests.post(f'{API}/gradebooks/import', json={'data_base64': b64}, headers=H).json()
print('\n5. import round-trip:', imp)
gb3 = requests.get(f'{API}/gradebooks/{imp["gradebook_id"]}', headers=H).json()
print('   re-imported students:', len(gb3['students']), '| grade:', gb3['grade'], '| section:', gb3['section'])
q1_vals = [gb3['scores']['1'].get(s['id'],{}).get('q1') for s in gb3['students']]
print('   q1 scores after round-trip:', q1_vals)

# 8. import the USER'S ORIGINAL file
b64o = base64.b64encode(open('/tmp/gradebook.xlsx','rb').read()).decode()
impo = requests.post(f'{API}/gradebooks/import', json={'data_base64': b64o}, headers=H).json()
print('\n6. original official file import:', impo)
gbo = requests.get(f'{API}/gradebooks/{impo["gradebook_id"]}', headers=H).json()
print('   students:', len(gbo['students']), '| sample scores sem1 first student:', gbo['scores']['1'].get(gbo['students'][0]['id']))

# cleanup
requests.delete(f'{API}/quizzes/{qid}', headers=H)
for g in requests.get(f'{API}/gradebooks', headers=H).json():
    requests.delete(f'{API}/gradebooks/{g["id"]}', headers=H)
print('\nALL TESTS DONE')
