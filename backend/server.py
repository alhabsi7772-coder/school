from fastapi import FastAPI, APIRouter, HTTPException, Depends, File, UploadFile
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from contextlib import asynccontextmanager
import os
import logging
import json
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import random
import base64
from emergentintegrations.llm.chat import LlmChat, UserMessage

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'al-khairat-school-it-secret-key-2024!')
JWT_ALGORITHM = "HS256"
security = HTTPBearer()

# =================== ACADEMIC YEARS (5 سنوات — كل عام به فصلان) ===================
ACADEMIC_YEARS = ["2025-2026", "2026-2027", "2027-2028", "2028-2029", "2029-2030"]
DEFAULT_ACADEMIC_YEAR = "2025-2026"
DEFAULT_SEMESTER = "1"


def teacher_year(t) -> str:
    """العام الدراسي النشط للمعلم الحالي"""
    return t["teacher"].get("academic_year", DEFAULT_ACADEMIC_YEAR)


def teacher_semester(t) -> str:
    """الفصل الدراسي النشط للمعلم الحالي (تفضيل عرض فقط — لا يؤثر على البيانات)"""
    return t["teacher"].get("semester", DEFAULT_SEMESTER)


def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()


def verify_password(p: str, h: str) -> bool:
    return bcrypt.checkpw(p.encode(), h.encode())


def make_token(data: dict) -> str:
    payload = {**data, "exp": datetime.now(timezone.utc) + timedelta(hours=168)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_teacher(creds: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="انتهت صلاحية الجلسة")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="جلسة غير صالحة")
    tid = payload.get("teacher_id")
    if not tid:
        raise HTTPException(status_code=401, detail="انتهت صلاحية الجلسة، يرجى تسجيل الدخول من جديد")
    teacher = await db.teachers.find_one({"id": tid}, {"_id": 0, "password_hash": 0})
    if not teacher:
        raise HTTPException(status_code=401, detail="الحساب غير موجود")
    if not teacher.get("is_active", True):
        raise HTTPException(status_code=403, detail="تم تعطيل هذا الحساب، تواصل مع مدير المنصة")
    return {"teacher_id": tid, "role": teacher.get("role", "teacher"), "teacher": teacher}


async def require_admin(t=Depends(get_teacher)):
    if t["role"] != "admin":
        raise HTTPException(status_code=403, detail="هذه الصلاحية لمدير المنصة فقط")
    return t


def now_iso():
    return datetime.now(timezone.utc).isoformat()


async def init_db():
    await db.teachers.create_index("username", unique=True, sparse=True)
    await db.login_attempts.create_index("identifier")

    # ترقية الحساب القديم (بدون اسم مستخدم) إلى حساب المدير
    legacy = await db.teachers.find_one({"username": {"$exists": False}})
    if legacy:
        await db.teachers.update_one({"id": legacy["id"]}, {"$set": {
            "username": "admin", "role": "admin", "is_active": True
        }})

    admin = await db.teachers.find_one({"role": "admin"})
    if not admin:
        admin = {
            "id": str(uuid.uuid4()),
            "username": "admin",
            "password_hash": hash_password("teacher123"),
            "school_name": "مدرسة الخيرات للتعليم الأساسي",
            "teacher_name": "أستاذ تقنية المعلومات",
            "role": "admin",
            "is_active": True,
            "created_at": now_iso()
        }
        await db.teachers.insert_one(admin)
    admin_id = admin["id"]

    # إنشاء 3 حسابات معلمين افتراضية
    seed_teachers = [
        ("teacher1", "khairat1", "المعلم الأول"),
        ("teacher2", "khairat2", "المعلم الثاني"),
        ("teacher3", "khairat3", "المعلم الثالث"),
    ]
    for username, pwd, name in seed_teachers:
        if not await db.teachers.find_one({"username": username}):
            await db.teachers.insert_one({
                "id": str(uuid.uuid4()),
                "username": username,
                "password_hash": hash_password(pwd),
                "school_name": "مدرسة الخيرات للتعليم الأساسي",
                "teacher_name": name,
                "role": "teacher",
                "is_active": True,
                "created_at": now_iso()
            })

    # ربط البيانات القديمة (بدون مالك) بحساب المدير
    await db.quizzes.update_many({"owner_id": {"$exists": False}}, {"$set": {"owner_id": admin_id}})
    await db.projects.update_many({"owner_id": {"$exists": False}}, {"$set": {"owner_id": admin_id}})
    await db.question_bank.update_many({"owner_id": {"$exists": False}}, {"$set": {"owner_id": admin_id}})

    # العام الدراسي: ترحيل البيانات القديمة إلى العام الافتراضي 2025/2026
    await db.teachers.update_many({"academic_year": {"$exists": False}}, {"$set": {"academic_year": DEFAULT_ACADEMIC_YEAR}})
    await db.teachers.update_many({"semester": {"$exists": False}}, {"$set": {"semester": DEFAULT_SEMESTER}})
    for coll in (db.quizzes, db.projects, db.gradebooks, db.rubrics):
        await coll.update_many({"year": {"$exists": False}}, {"$set": {"year": DEFAULT_ACADEMIC_YEAR}})

    # الفصل الدراسي للاختبارات والمشاريع: بيانات المدير القديمة → الفصل الثاني (بطلب المستخدم)، والبقية → الفصل الأول
    for coll in (db.quizzes, db.projects):
        await coll.update_many({"semester": {"$exists": False}, "owner_id": admin_id}, {"$set": {"semester": "2"}})
        await coll.update_many({"semester": {"$exists": False}}, {"$set": {"semester": "1"}})


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    yield
    client.close()


app = FastAPI(lifespan=lifespan)
api_router = APIRouter(prefix="/api")

# =================== MODELS ===================

class LoginReq(BaseModel):
    username: str
    password: str


class ChangePwdReq(BaseModel):
    old_password: str
    new_password: str


class TeacherProfileUpdate(BaseModel):
    teacher_name: Optional[str] = None
    school_name: Optional[str] = None


class QuizSettings(BaseModel):
    time_limit: Optional[int] = None
    secret_code: str = ""
    show_results: bool = True
    home_exam: bool = False
    randomize_questions: bool = True
    question_count: Optional[int] = None
    show_question_nav: bool = False


class QuizCreate(BaseModel):
    title: str
    description: str = ""
    settings: QuizSettings = Field(default_factory=QuizSettings)


class QuizUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    settings: Optional[QuizSettings] = None


class QuestionCreate(BaseModel):
    type: str
    text: str
    image_url: Optional[str] = None
    options: Optional[List[str]] = None
    correct_answer: Optional[str] = None
    points: float = 1.0


class QuestionUpdate(BaseModel):
    type: Optional[str] = None
    text: Optional[str] = None
    image_url: Optional[str] = None
    options: Optional[List[str]] = None
    correct_answer: Optional[str] = None
    points: Optional[float] = None


class StudentJoinReq(BaseModel):
    student_name: str
    grade: str
    section: str
    civil_id: Optional[str] = None


class AnswerItem(BaseModel):
    question_id: str
    answer_text: str


class SubmitReq(BaseModel):
    answers: List[AnswerItem]


class GradeItem(BaseModel):
    question_id: str
    score: float
    is_correct: bool


# =================== PROJECT MODELS ===================

class ProjectCreate(BaseModel):
    title: str
    description: str = ""
    deadline: Optional[str] = None

class ProjectFileInput(BaseModel):
    filename: str
    content_type: str
    data_base64: str
    size_bytes: int

class ProjectSubmitReq(BaseModel):
    student_name: str
    grade: str
    section: str
    files: List[ProjectFileInput]


# =================== AUTH ROUTES ===================

MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_MINUTES = 15


@api_router.post("/auth/login")
async def login(req: LoginReq):
    username = req.username.strip().lower()
    att = await db.login_attempts.find_one({"identifier": username})
    if att and att.get("locked_until"):
        if datetime.fromisoformat(att["locked_until"]) > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail=f"تم قفل تسجيل الدخول مؤقتاً بسبب محاولات خاطئة متكررة، حاول بعد {LOCKOUT_MINUTES} دقيقة")

    doc = await db.teachers.find_one({"username": username})
    if not doc or not verify_password(req.password, doc['password_hash']):
        count = (att or {}).get("count", 0) + 1
        upd = {"count": count, "locked_until": None}
        if count >= MAX_LOGIN_ATTEMPTS:
            upd["locked_until"] = (datetime.now(timezone.utc) + timedelta(minutes=LOCKOUT_MINUTES)).isoformat()
            upd["count"] = 0
        await db.login_attempts.update_one({"identifier": username}, {"$set": upd}, upsert=True)
        raise HTTPException(status_code=401, detail="اسم المستخدم أو كلمة المرور غير صحيحة")
    if not doc.get("is_active", True):
        raise HTTPException(status_code=403, detail="تم تعطيل هذا الحساب، تواصل مع مدير المنصة")

    await db.login_attempts.delete_one({"identifier": username})
    role = doc.get("role", "teacher")
    return {
        "token": make_token({"role": role, "teacher_id": doc["id"], "username": username}),
        "teacher_name": doc.get('teacher_name', 'المعلم'),
        "role": role,
        "username": username,
        "academic_year": doc.get("academic_year", DEFAULT_ACADEMIC_YEAR),
        "semester": doc.get("semester", DEFAULT_SEMESTER)
    }


@api_router.post("/auth/change-password")
async def change_password(req: ChangePwdReq, t=Depends(get_teacher)):
    doc = await db.teachers.find_one({"id": t["teacher_id"]})
    if not verify_password(req.old_password, doc['password_hash']):
        raise HTTPException(status_code=400, detail="كلمة المرور القديمة غير صحيحة")
    await db.teachers.update_one({"id": t["teacher_id"]}, {"$set": {"password_hash": hash_password(req.new_password)}})
    return {"message": "تم تغيير كلمة المرور"}


@api_router.get("/auth/profile")
async def get_profile(t=Depends(get_teacher)):
    doc = await db.teachers.find_one({"id": t["teacher_id"]}, {"_id": 0, "password_hash": 0})
    return doc


@api_router.put("/auth/profile")
async def update_profile(data: TeacherProfileUpdate, t=Depends(get_teacher)):
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    await db.teachers.update_one({"id": t["teacher_id"]}, {"$set": update})
    return {"message": "تم التحديث"}


# =================== APP SETTINGS ===================

@api_router.get("/quizzes/{qid}/monitor")
async def get_quiz_monitor_data(qid: str, t=Depends(get_teacher)):
    """Combined endpoint: quiz info + lobby list in a single DB round-trip."""
    q = await db.quizzes.find_one({"id": qid, "owner_id": t["teacher_id"]}, {"_id": 0, "questions": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    subs = await db.submissions.find(
        {"quiz_id": qid},
        {"_id": 0, "id": 1, "student_name": 1, "grade": 1, "section": 1,
         "score": 1, "submitted": 1, "start_time": 1, "total_score": 1, "max_score": 1}
    ).to_list(500)
    return {"quiz": q, "submissions": subs}


@api_router.get("/app-settings")
async def get_app_settings():
    settings = await db.app_settings.find_one({}, {"_id": 0}) or {}
    return {"student_mode": settings.get("student_mode", "dark")}


# =================== ACADEMIC YEAR ROUTES ===================

class AcademicYearReq(BaseModel):
    year: str


class SemesterReq(BaseModel):
    semester: str


@api_router.get("/academic-years")
async def get_academic_years(t=Depends(get_teacher)):
    """قائمة الأعوام الدراسية المتاحة (5 سنوات) + العام والفصل النشطان للمعلم الحالي"""
    return {"years": ACADEMIC_YEARS, "current": teacher_year(t), "semester": teacher_semester(t)}


@api_router.put("/academic-years/current")
async def set_academic_year(data: AcademicYearReq, t=Depends(get_teacher)):
    """تبديل العام الدراسي النشط — خاص بكل معلم على حدة"""
    if data.year not in ACADEMIC_YEARS:
        raise HTTPException(400, "العام الدراسي غير صحيح")
    await db.teachers.update_one({"id": t["teacher_id"]}, {"$set": {"academic_year": data.year}})
    return {"message": "تم تغيير العام الدراسي", "current": data.year}


@api_router.put("/academic-years/semester")
async def set_semester(data: SemesterReq, t=Depends(get_teacher)):
    """تبديل الفصل الدراسي النشط — تفضيل عرض فقط، لا يغيّر أي بيانات أو درجات"""
    if data.semester not in ("1", "2"):
        raise HTTPException(400, "الفصل الدراسي غير صحيح")
    await db.teachers.update_one({"id": t["teacher_id"]}, {"$set": {"semester": data.semester}})
    return {"message": "تم تغيير الفصل الدراسي", "semester": data.semester}


@api_router.put("/app-settings")
async def update_app_settings(data: dict, _=Depends(get_teacher)):
    await db.app_settings.update_one({}, {"$set": data}, upsert=True)
    return {"success": True}


# =================== QUIZ ROUTES ===================

@api_router.get("/quizzes")
async def list_quizzes(t=Depends(get_teacher)):
    pipeline = [
        {"$match": {"owner_id": t["teacher_id"], "year": teacher_year(t), "semester": teacher_semester(t)}},
        {"$lookup": {
            "from": "submissions",
            "let": {"quiz_id": "$id"},
            "pipeline": [
                {"$match": {"$expr": {"$eq": ["$quiz_id", "$$quiz_id"]}, "submitted_at": {"$ne": None}}},
                {"$count": "count"}
            ],
            "as": "sub_data"
        }},
        {"$addFields": {
            "submission_count": {"$ifNull": [{"$arrayElemAt": ["$sub_data.count", 0]}, 0]},
            "question_count": {"$size": {"$ifNull": ["$questions", []]}}
        }},
        {"$project": {"_id": 0, "sub_data": 0}}
    ]
    quizzes = await db.quizzes.aggregate(pipeline).to_list(1000)
    return sorted(quizzes, key=lambda x: x.get('created_at', ''), reverse=True)


@api_router.post("/quizzes")
async def create_quiz(data: QuizCreate, t=Depends(get_teacher)):
    quiz = {
        "id": str(uuid.uuid4()),
        "owner_id": t["teacher_id"],
        "year": teacher_year(t),
        "semester": teacher_semester(t),
        "title": data.title,
        "description": data.description,
        "questions": [],
        "settings": data.settings.model_dump(),
        "status": "draft",
        "start_time": None,
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    await db.quizzes.insert_one(quiz)
    quiz.pop('_id', None)
    return quiz


@api_router.get("/quizzes/{qid}")
async def get_quiz(qid: str, t=Depends(get_teacher)):
    q = await db.quizzes.find_one({"id": qid, "owner_id": t["teacher_id"]}, {"_id": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    return q


@api_router.put("/quizzes/{qid}")
async def update_quiz(qid: str, data: QuizUpdate, t=Depends(get_teacher)):
    upd = {k: v for k, v in data.model_dump().items() if v is not None}
    if 'settings' in upd and data.settings:
        upd['settings'] = data.settings.model_dump()
    upd['updated_at'] = now_iso()
    r = await db.quizzes.update_one({"id": qid, "owner_id": t["teacher_id"]}, {"$set": upd})
    if r.matched_count == 0:
        raise HTTPException(404, "الاختبار غير موجود")
    return {"message": "تم التحديث"}


@api_router.delete("/quizzes/{qid}")
async def delete_quiz(qid: str, t=Depends(get_teacher)):
    r = await db.quizzes.delete_one({"id": qid, "owner_id": t["teacher_id"]})
    if r.deleted_count == 0:
        raise HTTPException(404, "الاختبار غير موجود")
    await db.submissions.delete_many({"quiz_id": qid})
    return {"message": "تم الحذف"}


@api_router.post("/quizzes/{qid}/duplicate")
async def duplicate_quiz(qid: str, t=Depends(get_teacher)):
    q = await db.quizzes.find_one({"id": qid, "owner_id": t["teacher_id"]}, {"_id": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    import string
    new_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    new_questions = []
    for qq in q.get('questions', []):
        new_q = dict(qq)
        new_q['id'] = str(uuid.uuid4())
        new_questions.append(new_q)
    new_settings = dict(q.get('settings', {}))
    new_settings['secret_code'] = new_code
    new_quiz = {
        "id": str(uuid.uuid4()),
        "owner_id": t["teacher_id"],
        "year": teacher_year(t),
        "semester": teacher_semester(t),
        "title": q['title'] + " (نسخة)",
        "description": q.get('description', ''),
        "questions": new_questions,
        "settings": new_settings,
        "status": "draft",
        "start_time": None,
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    await db.quizzes.insert_one(new_quiz)
    new_quiz.pop('_id', None)
    return new_quiz


@api_router.post("/quizzes/{qid}/activate")
async def activate(qid: str, t=Depends(get_teacher)):
    await db.quizzes.update_one({"id": qid, "owner_id": t["teacher_id"]}, {"$set": {"status": "waiting", "updated_at": now_iso()}})
    return {"message": "تم التفعيل"}


@api_router.post("/quizzes/{qid}/start")
async def start_quiz(qid: str, t=Depends(get_teacher)):
    await db.quizzes.update_one({"id": qid, "owner_id": t["teacher_id"]}, {"$set": {"status": "active", "start_time": now_iso(), "updated_at": now_iso()}})
    return {"message": "بدأ الاختبار"}


@api_router.post("/quizzes/{qid}/close")
async def close_quiz(qid: str, t=Depends(get_teacher)):
    await db.quizzes.update_one({"id": qid, "owner_id": t["teacher_id"]}, {"$set": {"status": "closed", "updated_at": now_iso()}})
    return {"message": "تم الإغلاق"}


@api_router.post("/quizzes/{qid}/reset")
async def reset_quiz(qid: str, t=Depends(get_teacher)):
    r = await db.quizzes.update_one({"id": qid, "owner_id": t["teacher_id"]}, {"$set": {"status": "draft", "start_time": None, "updated_at": now_iso()}})
    if r.matched_count == 0:
        raise HTTPException(404, "الاختبار غير موجود")
    await db.submissions.delete_many({"quiz_id": qid})
    return {"message": "تم إعادة الضبط"}


# =================== QUESTIONS ===================

@api_router.post("/quizzes/{qid}/questions")
async def add_question(qid: str, data: QuestionCreate, t=Depends(get_teacher)):
    q = await db.quizzes.find_one({"id": qid, "owner_id": t["teacher_id"]})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    question = {
        "id": str(uuid.uuid4()),
        "type": data.type,
        "text": data.text,
        "image_url": data.image_url,
        "options": data.options,
        "correct_answer": data.correct_answer,
        "points": data.points,
        "order": len(q.get('questions', []))
    }
    await db.quizzes.update_one(
        {"id": qid},
        {"$push": {"questions": question}, "$set": {"updated_at": now_iso()}}
    )
    return question


@api_router.put("/quizzes/{qid}/questions/{question_id}")
async def update_question(qid: str, question_id: str, data: QuestionUpdate, t=Depends(get_teacher)):
    upd = {f"questions.$.{k}": v for k, v in data.model_dump().items() if v is not None}
    upd['updated_at'] = now_iso()
    await db.quizzes.update_one({"id": qid, "owner_id": t["teacher_id"], "questions.id": question_id}, {"$set": upd})
    return {"message": "تم التحديث"}


@api_router.delete("/quizzes/{qid}/questions/{question_id}")
async def delete_question(qid: str, question_id: str, t=Depends(get_teacher)):
    await db.quizzes.update_one({"id": qid, "owner_id": t["teacher_id"]}, {"$pull": {"questions": {"id": question_id}}})
    return {"message": "تم الحذف"}


@api_router.post("/upload-image")
async def upload_image(file: UploadFile = File(...), _=Depends(get_teacher)):
    content = await file.read()
    if len(content) > 3 * 1024 * 1024:
        raise HTTPException(400, "الصورة كبيرة جداً (الحد الأقصى 3MB)")
    b64 = base64.b64encode(content).decode()
    return {"image_url": f"data:{file.content_type};base64,{b64}"}


# =================== STUDENT PUBLIC ROUTES ===================

@api_router.get("/quiz/join/{code}")
async def check_code(code: str):
    q = await db.quizzes.find_one(
        {"settings.secret_code": code, "status": {"$in": ["waiting", "active"]}},
        {"_id": 0, "questions": 0}
    )
    if not q:
        raise HTTPException(404, "رمز الاختبار غير صحيح أو الاختبار غير متاح")
    return {
        "id": q['id'],
        "title": q['title'],
        "description": q['description'],
        "status": q['status'],
        "settings": {
            "time_limit": q['settings'].get('time_limit'),
            "home_exam": q['settings'].get('home_exam', False),
            "show_results": q['settings'].get('show_results', True)
        }
    }


@api_router.post("/quiz/{qid}/join")
async def join_quiz(qid: str, data: StudentJoinReq):
    q = await db.quizzes.find_one({"id": qid, "status": {"$in": ["waiting", "active"]}}, {"_id": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير متاح")
    if q['settings'].get('home_exam') and not data.civil_id:
        raise HTTPException(400, "الرقم المدني مطلوب لهذا الاختبار")

    existing = await db.submissions.find_one({
        "quiz_id": qid, "student_name": data.student_name,
        "grade": data.grade, "section": data.section
    }, {"_id": 0})
    if existing:
        return {"submission_id": existing['id'], "already_joined": True, "status": q['status']}

    sub = {
        "id": str(uuid.uuid4()), "quiz_id": qid,
        "student_name": data.student_name, "grade": data.grade,
        "section": data.section, "civil_id": data.civil_id,
        "answers": [], "question_order": [],
        "total_score": 0, "max_score": 0, "percentage": 0,
        "is_graded": False, "submitted_at": None, "started_at": now_iso()
    }
    await db.submissions.insert_one(sub)
    return {"submission_id": sub['id'], "already_joined": False, "status": q['status']}


@api_router.get("/quiz/{qid}/status")
async def quiz_status(qid: str):
    q = await db.quizzes.find_one({"id": qid}, {"_id": 0, "questions": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    return {"status": q['status'], "start_time": q.get('start_time'), "time_limit": q['settings'].get('time_limit')}


@api_router.get("/quiz/{qid}/questions/{sub_id}")
async def get_questions(qid: str, sub_id: str):
    q = await db.quizzes.find_one({"id": qid}, {"_id": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    if q['status'] != 'active':
        return {"status": q['status'], "questions": [], "time_limit": None, "start_time": None}

    sub = await db.submissions.find_one({"id": sub_id, "quiz_id": qid})
    if not sub:
        raise HTTPException(404, "لم يتم العثور على الطالب")

    questions = q.get('questions', [])
    if sub.get('question_order') and len(sub['question_order']) > 0:
        q_map = {qq['id']: qq for qq in questions}
        questions = [q_map[qid2] for qid2 in sub['question_order'] if qid2 in q_map]
    else:
        if q['settings'].get('randomize_questions', True):
            random.shuffle(questions)
        cnt = q['settings'].get('question_count')
        if cnt and cnt < len(questions):
            questions = questions[:cnt]
        order = [qq['id'] for qq in questions]
        await db.submissions.update_one({"id": sub_id}, {"$set": {"question_order": order}})

    safe = []
    for qq in questions:
        item = {"id": qq['id'], "type": qq['type'], "text": qq['text'],
                "image_url": qq.get('image_url'), "points": qq.get('points', 1)}
        if qq['type'] == 'mcq':
            item['options'] = qq.get('options', [])
        elif qq['type'] == 'true_false':
            item['options'] = ['صح', 'خطأ']
        safe.append(item)

    return {"status": "active", "questions": safe,
            "time_limit": q['settings'].get('time_limit'), "start_time": q.get('start_time'),
            "show_question_nav": q['settings'].get('show_question_nav', False)}


@api_router.post("/quiz/{qid}/submit/{sub_id}")
async def submit(qid: str, sub_id: str, data: SubmitReq):
    q = await db.quizzes.find_one({"id": qid}, {"_id": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    sub = await db.submissions.find_one({"id": sub_id, "quiz_id": qid})
    if not sub:
        raise HTTPException(404, "لم يتم العثور على الطالب")

    q_map = {qq['id']: qq for qq in q.get('questions', [])}
    answers = []
    total = 0
    max_score = 0
    has_long = False

    for ans in data.answers:
        qq = q_map.get(ans.question_id)
        if not qq:
            continue
        pts = qq.get('points', 1)
        max_score += pts
        rec = {"question_id": ans.question_id, "answer_text": ans.answer_text,
               "is_correct": None, "score": 0, "manual_score": None}

        if qq['type'] in ['mcq', 'true_false']:
            correct = (qq.get('correct_answer') or '').strip().lower()
            student = ans.answer_text.strip().lower()
            rec['is_correct'] = correct == student
            rec['score'] = pts if rec['is_correct'] else 0
            total += rec['score']
        elif qq['type'] == 'short':
            correct = (qq.get('correct_answer') or '').strip()
            student = ans.answer_text.strip()
            rec['is_correct'] = correct.lower() == student.lower()
            rec['score'] = pts if rec['is_correct'] else 0
            total += rec['score']
        elif qq['type'] == 'long':
            has_long = True
        answers.append(rec)

    is_graded = not has_long
    pct = (total / max_score * 100) if max_score > 0 else 0

    await db.submissions.update_one({"id": sub_id}, {"$set": {
        "answers": answers, "total_score": total, "max_score": max_score,
        "percentage": pct, "is_graded": is_graded, "submitted_at": now_iso()
    }})
    return {"message": "تم تسليم الاختبار", "is_graded": is_graded,
            "show_results": q['settings'].get('show_results', True)}


@api_router.get("/quiz/{qid}/result/{sub_id}")
async def get_result(qid: str, sub_id: str):
    q = await db.quizzes.find_one({"id": qid}, {"_id": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    sub = await db.submissions.find_one({"id": sub_id, "quiz_id": qid}, {"_id": 0})
    if not sub:
        raise HTTPException(404, "لم يتم العثور على الطالب")

    if not sub.get('submitted_at'):
        return {"submitted": False}

    if not q['settings'].get('show_results', True) and not sub.get('is_graded'):
        return {"show_results": False, "submitted": True, "is_graded": sub.get('is_graded', False)}

    q_map = {qq['id']: qq for qq in q.get('questions', [])}
    enriched = []
    for ans in sub.get('answers', []):
        qq = q_map.get(ans['question_id'], {})
        enriched.append({
            **ans,
            "question_text": qq.get('text', ''),
            "question_type": qq.get('type', ''),
            "correct_answer": qq.get('correct_answer') if qq.get('type') != 'long' else None,
            "image_url": qq.get('image_url'),
            "points": qq.get('points', 1)
        })

    return {
        "show_results": True, "submitted": True,
        "student_name": sub['student_name'], "grade": sub['grade'],
        "section": sub['section'], "total_score": sub['total_score'],
        "max_score": sub['max_score'], "percentage": sub['percentage'],
        "is_graded": sub['is_graded'], "answers": enriched, "quiz_title": q['title']
    }


# =================== TEACHER RESULTS ===================

@api_router.get("/quizzes/{qid}/lobby")
async def get_lobby(qid: str, t=Depends(get_teacher)):
    q = await db.quizzes.find_one({"id": qid, "owner_id": t["teacher_id"]}, {"id": 1})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    subs = await db.submissions.find(
        {"quiz_id": qid},
        {"_id": 0, "student_name": 1, "grade": 1, "section": 1, "started_at": 1, "submitted_at": 1, "id": 1}
    ).to_list(1000)
    return subs


@api_router.get("/quizzes/{qid}/results")
async def quiz_results(qid: str, t=Depends(get_teacher)):
    q = await db.quizzes.find_one({"id": qid, "owner_id": t["teacher_id"]}, {"_id": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    subs = await db.submissions.find(
        {"quiz_id": qid, "submitted_at": {"$ne": None}}, {"_id": 0}
    ).to_list(1000)
    scores = [s.get('percentage', 0) for s in subs]
    return {
        "quiz": q, "submissions": subs,
        "stats": {
            "total": len(subs),
            "avg": round(sum(scores) / len(scores), 1) if scores else 0,
            "max": max(scores, default=0),
            "min": min(scores, default=0)
        }
    }


@api_router.get("/quizzes/{qid}/pending")
async def pending_grading(qid: str, t=Depends(get_teacher)):
    q = await db.quizzes.find_one({"id": qid, "owner_id": t["teacher_id"]}, {"_id": 0})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    subs = await db.submissions.find(
        {"quiz_id": qid, "is_graded": False, "submitted_at": {"$ne": None}}, {"_id": 0}
    ).to_list(1000)
    return {"submissions": subs, "quiz": q}


@api_router.post("/quizzes/{qid}/grade/{sub_id}")
async def grade_submission(qid: str, sub_id: str, grades: List[GradeItem], t=Depends(get_teacher)):
    quiz = await db.quizzes.find_one({"id": qid, "owner_id": t["teacher_id"]}, {"id": 1})
    if not quiz:
        raise HTTPException(404, "الاختبار غير موجود")
    sub = await db.submissions.find_one({"id": sub_id, "quiz_id": qid})
    if not sub:
        raise HTTPException(404, "لم يتم العثور على الطالب")

    g_map = {g.question_id: g for g in grades}
    answers = sub.get('answers', [])
    total = 0

    for ans in answers:
        g = g_map.get(ans['question_id'])
        if g:
            ans['score'] = g.score
            ans['manual_score'] = g.score
            ans['is_correct'] = g.is_correct
        total += ans.get('score', 0)

    max_s = sub.get('max_score', 0)
    pct = (total / max_s * 100) if max_s > 0 else 0

    await db.submissions.update_one({"id": sub_id}, {"$set": {
        "answers": answers, "total_score": total, "percentage": pct, "is_graded": True
    }})
    return {"message": "تم التصحيح", "total_score": total, "percentage": pct}


@api_router.put("/quizzes/{qid}/toggle-results")
async def toggle_results(qid: str, t=Depends(get_teacher)):
    q = await db.quizzes.find_one({"id": qid, "owner_id": t["teacher_id"]})
    if not q:
        raise HTTPException(404, "الاختبار غير موجود")
    current = q.get('settings', {}).get('show_results', True)
    await db.quizzes.update_one({"id": qid, "owner_id": t["teacher_id"]}, {"$set": {"settings.show_results": not current}})
    return {"show_results": not current}


# =================== PROJECT ROUTES ===================
import string as _string

@api_router.get("/projects")
async def list_projects(t=Depends(get_teacher)):
    pipeline = [
        {"$match": {"owner_id": t["teacher_id"], "year": teacher_year(t), "semester": teacher_semester(t)}},
        {"$lookup": {
            "from": "project_submissions",
            "let": {"project_id": "$id"},
            "pipeline": [
                {"$match": {"$expr": {"$eq": ["$project_id", "$$project_id"]}}},
                {"$count": "count"}
            ],
            "as": "sub_data"
        }},
        {"$addFields": {
            "submission_count": {"$ifNull": [{"$arrayElemAt": ["$sub_data.count", 0]}, 0]}
        }},
        {"$project": {"_id": 0, "sub_data": 0}}
    ]
    projects = await db.projects.aggregate(pipeline).to_list(1000)
    return sorted(projects, key=lambda x: x.get('created_at', ''), reverse=True)


@api_router.post("/projects")
async def create_project(data: ProjectCreate, t=Depends(get_teacher)):
    code = ''.join(random.choices(_string.ascii_uppercase + _string.digits, k=6))
    project = {
        "id": str(uuid.uuid4()),
        "owner_id": t["teacher_id"],
        "year": teacher_year(t),
        "semester": teacher_semester(t),
        "title": data.title,
        "description": data.description,
        "deadline": data.deadline,
        "code": code,
        "created_at": now_iso()
    }
    await db.projects.insert_one(project)
    project.pop('_id', None)
    return project


@api_router.delete("/projects/{pid}")
async def delete_project(pid: str, _=Depends(get_teacher)):
    subs = await db.project_submissions.find({"project_id": pid}, {"_id": 0, "id": 1}).to_list(1000)
    for s in subs:
        await db.project_files.delete_many({"submission_id": s['id']})
    await db.project_submissions.delete_many({"project_id": pid})
    await db.projects.delete_one({"id": pid})
    return {"message": "تم الحذف"}


@api_router.get("/project/{code}/info")
async def project_info(code: str):
    p = await db.projects.find_one({"code": code}, {"_id": 0})
    if not p:
        raise HTTPException(404, "المشروع غير موجود")
    p.pop('_id', None)
    return p


@api_router.post("/project/{code}/submit")
async def submit_project(code: str, data: ProjectSubmitReq):
    p = await db.projects.find_one({"code": code}, {"_id": 0})
    if not p:
        raise HTTPException(404, "المشروع غير موجود")
    if not data.files:
        raise HTTPException(400, "يجب إرفاق ملف واحد على الأقل")
    if len(data.files) > 15:
        raise HTTPException(400, "الحد الأقصى 15 ملف")
    for f in data.files:
        if f.size_bytes > 20 * 1024 * 1024:
            raise HTTPException(400, f"حجم الملف {f.filename} يتجاوز 20MB")

    sub_id = str(uuid.uuid4())
    file_metas = []
    for f in data.files:
        fid = str(uuid.uuid4())
        await db.project_files.insert_one({
            "id": fid,
            "submission_id": sub_id,
            "filename": f.filename,
            "content_type": f.content_type,
            "size_bytes": f.size_bytes,
            "data_base64": f.data_base64
        })
        file_metas.append({"id": fid, "filename": f.filename, "content_type": f.content_type, "size_bytes": f.size_bytes})

    await db.project_submissions.insert_one({
        "id": sub_id,
        "project_id": p['id'],
        "student_name": data.student_name,
        "grade": data.grade,
        "section": data.section,
        "files": file_metas,
        "submitted_at": now_iso()
    })
    return {"message": "تم تسليم المشروع بنجاح", "submission_id": sub_id}


@api_router.get("/projects/{pid}/submissions")
async def project_submissions(pid: str, t=Depends(get_teacher)):
    p = await db.projects.find_one({"id": pid, "owner_id": t["teacher_id"]}, {"_id": 0})
    if not p:
        raise HTTPException(404, "المشروع غير موجود")
    subs = await db.project_submissions.find({"project_id": pid}, {"_id": 0}).to_list(1000)
    return {"project": p, "submissions": subs}


@api_router.get("/project-file/{file_id}")
async def download_project_file(file_id: str, _=Depends(get_teacher)):
    f = await db.project_files.find_one({"id": file_id}, {"_id": 0})
    if not f:
        raise HTTPException(404, "الملف غير موجود")
    return f


# =================== QUESTION BANK MODELS ===================

class GenerateQuestionsReq(BaseModel):
    grade: str  # "5" or "8"
    topic: Optional[str] = None

class CreateQuizFromBankReq(BaseModel):
    title: str
    question_ids: List[str]


# =================== QUESTION BANK ROUTES ===================

GRADE5_TOPICS = [
    "الجداول الإلكترونية (Microsoft Excel) - إدخال البيانات والصيغ الحسابية",
    "الجداول الإلكترونية - دوال SUM وAVERAGE وMAX وMIN وCOUNT وIF",
    "الجداول الإلكترونية - التحقق من صحة البيانات والتنسيق الشرطي",
    "الجداول الإلكترونية - الرسوم البيانية وأنواعها",
    "الشبكات والإنترنت - أنواع الشبكات LAN وWAN",
    "خدمات الإنترنت والبريد الإلكتروني",
    "الأمن المعلوماتي وحماية الخصوصية"
]

GRADE8_TOPICS = [
    "قواعد البيانات - المفاهيم الأساسية: الجداول والحقول والسجلات",
    "Microsoft Access - إنشاء الجداول والعلاقات بين الجداول",
    "Microsoft Access - الاستعلامات (Queries) وأنواعها",
    "Microsoft Access - النماذج (Forms) والتقارير (Reports)",
    "تطوير الويب - بنية صفحة HTML والعلامات الأساسية",
    "تطوير الويب - تنسيق CSS والعناصر الأساسية",
    "الأمن المعلوماتي - التهديدات الإلكترونية وأساليب الحماية"
]

def build_generation_prompt(grade: str, topic: str) -> str:
    return f"""أنت أستاذ متخصص في تقنية المعلومات في سلطنة عُمان. مهمتك هي إنشاء أسئلة امتحانية دقيقة وصحيحة لمادة تقنية المعلومات.

قم بإنشاء بالضبط 15 سؤالاً للصف {grade} - الفصل الثاني حول الموضوع التالي:
{topic}

المتطلبات:
- الأسئلة باللغة العربية فقط (يُسمح بالمصطلحات التقنية الإنجليزية مثل: Excel, Access, HTML, CSS)
- التنوع في الأنواع: 6 أسئلة اختيارية (mcq) + 4 صح/خطأ (true_false) + 3 إجابة قصيرة (short) + 2 إجابة طويلة (long)
- التنوع في الصعوبة: 5 سهلة (easy) + 6 متوسطة (medium) + 4 صعبة (hard)
- كل سؤال يجب أن يكون صحيحاً علمياً ومناسباً لمستوى الصف {grade}

أعد الإجابة بصيغة JSON فقط كمصفوفة من 15 سؤالاً. كل سؤال له الشكل التالي:
{{
  "text": "نص السؤال",
  "type": "mcq|true_false|short|long",
  "options": ["خيار أ", "خيار ب", "خيار ج", "خيار د"],  // للأسئلة الاختيارية فقط، null لغيرها
  "correct_answer": "الإجابة الصحيحة",  // للاختيارية: نص الخيار الصحيح، للصح/خطأ: "صح" أو "خطأ"، للقصيرة: الإجابة النموذجية، للطويلة: null
  "points": 1,  // الاختيارية: 1، الصح/خطأ: 1، القصيرة: 2، الطويلة: 3
  "difficulty": "easy|medium|hard",
  "topic": "{topic}",
  "cognitive_level": "recall|understanding|application|analysis"
}}

مهم جداً: أعد JSON فقط بدون أي شرح أو نص إضافي."""


@api_router.post("/question-bank/generate")
async def generate_questions(req: GenerateQuestionsReq, t=Depends(get_teacher)):
    if req.grade not in ["5", "8"]:
        raise HTTPException(400, "الصف يجب أن يكون 5 أو 8")

    topics_list = GRADE5_TOPICS if req.grade == "5" else GRADE8_TOPICS
    topic = req.topic if req.topic else random.choice(topics_list)

    llm_key = os.environ.get('EMERGENT_LLM_KEY')
    if not llm_key:
        raise HTTPException(500, "مفتاح LLM غير مهيأ")

    chat = LlmChat(
        api_key=llm_key,
        session_id=str(uuid.uuid4()),
        system_message="أنت أستاذ متخصص في تقنية المعلومات. تُنشئ أسئلة امتحانية بصيغة JSON فقط."
    ).with_model("gemini", "gemini-2.0-flash")

    prompt = build_generation_prompt(req.grade, topic)
    user_msg = UserMessage(text=prompt)

    try:
        response_text = await chat.send_message(user_msg)
    except Exception as e:
        raise HTTPException(500, f"خطأ في توليد الأسئلة: {str(e)}")

    # استخراج JSON من الاستجابة
    try:
        text = response_text.strip()
        if text.startswith("```"):
            lines = text.split("\n")
            text = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
        questions_data = json.loads(text)
        if not isinstance(questions_data, list):
            raise ValueError("الاستجابة ليست قائمة")
    except Exception as e:
        raise HTTPException(500, f"خطأ في تحليل استجابة الذكاء الاصطناعي: {str(e)}")

    saved = []
    for q in questions_data[:15]:
        item = {
            "id": str(uuid.uuid4()),
            "owner_id": t["teacher_id"],
            "text": q.get("text", ""),
            "type": q.get("type", "mcq"),
            "options": q.get("options") if q.get("type") == "mcq" else None,
            "correct_answer": q.get("correct_answer"),
            "points": float(q.get("points", 1)),
            "difficulty": q.get("difficulty", "medium"),
            "grade": req.grade,
            "topic": q.get("topic", topic),
            "cognitive_level": q.get("cognitive_level", "recall"),
            "created_at": now_iso()
        }
        await db.question_bank.insert_one(item)
        item.pop("_id", None)
        saved.append(item)

    return {"questions": saved, "count": len(saved), "topic": topic}


@api_router.get("/question-bank")
async def list_question_bank(
    grade: Optional[str] = None,
    difficulty: Optional[str] = None,
    type: Optional[str] = None,
    t=Depends(get_teacher)
):
    query = {"owner_id": t["teacher_id"]}
    if grade:
        query["grade"] = grade
    if difficulty:
        query["difficulty"] = difficulty
    if type:
        query["type"] = type

    questions = await db.question_bank.find(query, {"_id": 0}).to_list(2000)
    return sorted(questions, key=lambda x: x.get("created_at", ""), reverse=True)


@api_router.delete("/question-bank/{qid}")
async def delete_bank_question(qid: str, t=Depends(get_teacher)):
    result = await db.question_bank.delete_one({"id": qid, "owner_id": t["teacher_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "السؤال غير موجود")
    return {"message": "تم الحذف"}


@api_router.post("/question-bank/create-quiz")
async def create_quiz_from_bank(data: CreateQuizFromBankReq, t=Depends(get_teacher)):
    if not data.question_ids:
        raise HTTPException(400, "يجب اختيار سؤال واحد على الأقل")

    questions_cursor = await db.question_bank.find(
        {"id": {"$in": data.question_ids}, "owner_id": t["teacher_id"]}, {"_id": 0}
    ).to_list(1000)

    if not questions_cursor:
        raise HTTPException(404, "لم يتم العثور على الأسئلة المختارة")

    import string
    new_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

    quiz_questions = []
    for i, bq in enumerate(questions_cursor):
        quiz_questions.append({
            "id": str(uuid.uuid4()),
            "type": bq["type"],
            "text": bq["text"],
            "image_url": None,
            "options": bq.get("options"),
            "correct_answer": bq.get("correct_answer"),
            "points": bq.get("points", 1.0),
            "order": i
        })

    quiz = {
        "id": str(uuid.uuid4()),
        "owner_id": t["teacher_id"],
        "year": teacher_year(t),
        "semester": teacher_semester(t),
        "title": data.title,
        "description": "",
        "questions": quiz_questions,
        "settings": {
            "time_limit": None,
            "secret_code": new_code,
            "show_results": True,
            "home_exam": False,
            "randomize_questions": True,
            "question_count": None,
            "show_question_nav": False
        },
        "status": "draft",
        "start_time": None,
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    await db.quizzes.insert_one(quiz)
    quiz.pop("_id", None)
    return quiz


# =================== GRADEBOOK (سجل الدرجات) ===================
from io import BytesIO
from difflib import SequenceMatcher
import re as _re
import openpyxl as _xl
from openpyxl.styles import Font as _Font, Alignment as _Align, Border as _Border, Side as _Side, PatternFill as _Fill
from openpyxl.utils import get_column_letter as _col_letter
from fastapi import Response as _Response

GB_SEM1 = 'الفصل الأول'
GB_SEM2 = 'الفصل الثاني '
GB_FIELDS = [
    ("d1", "حوار 1", 10), ("d2", "حوار 2", 10),
    ("q1", "قصيرة 1", 5), ("q2", "قصيرة 2", 5), ("q3", "قصيرة 3", 5), ("q4", "قصيرة 4", 5),
    ("p1", "عملي 1", 20), ("p2", "عملي 2", 20),
    ("proj", "المشروع", 20),
]
GB_MAX = {k: m for k, _, m in GB_FIELDS}
GB_XL_COLS = [(3, "d1"), (4, "d2"), (6, "q1"), (7, "q2"), (8, "q3"), (9, "q4"), (11, "p1"), (12, "p2"), (14, "proj")]

# النموذج الرسمي للصفوف (7-10): الحوار 2×10 + الأنشطة العملية 2×20 + الاختبار القصير 20 + المشروع 20 = 100
GB_FIELDS_78 = [
    ("d1", "حوار 1", 10), ("d2", "حوار 2", 10),
    ("p1", "عملي 1", 20), ("p2", "عملي 2", 20),
    ("q1", "الاختبار القصير", 20),
    ("proj", "المشروع", 20),
]
GB_MAX_78 = {k: m for k, _, m in GB_FIELDS_78}
GB_XL_COLS_78 = [(3, "d1"), (4, "d2"), (6, "p1"), (7, "p2"), (9, "q1"), (10, "proj")]
GB_TEMPLATES = ("5-6", "7-10")


def gb_template(gb) -> str:
    return gb.get("template", "5-6")


def gb_max_map(gb) -> dict:
    return GB_MAX_78 if gb_template(gb) == "7-10" else GB_MAX


class GradebookCreate(BaseModel):
    grade: str
    section: str
    students: List[str] = []
    template: str = "5-6"

class GradebookStudentsAdd(BaseModel):
    names: List[str]
    position: Optional[int] = None

class GradebookReorder(BaseModel):
    student_ids: List[str]

class GradebookStudentRename(BaseModel):
    name: str

class GBScoreUpdate(BaseModel):
    semester: str
    student_id: str
    field: str
    value: Optional[float] = None

class GBScoresBulk(BaseModel):
    updates: List[GBScoreUpdate]

class GBMatchReq(BaseModel):
    quiz_id: str

class GBApplyItem(BaseModel):
    submission_id: str
    student_id: str

class GBApplyReq(BaseModel):
    quiz_id: str
    semester: str
    column: str
    mappings: List[GBApplyItem]

class GBImportReq(BaseModel):
    data_base64: str
    gradebook_id: Optional[str] = None


# ---------- Smart Arabic name matching ----------
AR_STOP = {"بن", "بنت", "ابن", "ابنة"}

def _norm_ar(s: str) -> str:
    s = _re.sub(r'[\u064B-\u0652\u0640]', '', s or '')
    for a, b in [('أ', 'ا'), ('إ', 'ا'), ('آ', 'ا'), ('ٱ', 'ا'), ('ة', 'ه'), ('ى', 'ي'), ('ؤ', 'و'), ('ئ', 'ي'), ('ء', '')]:
        s = s.replace(a, b)
    return _re.sub(r'\s+', ' ', s).strip().lower()

def _name_tokens(s: str):
    toks = [t for t in _norm_ar(s).split() if t and t not in AR_STOP]
    out = []
    for t in toks:
        if t.startswith('ال') and len(t) > 4:
            t = t[2:]
        out.append(t)
    return out

def name_similarity(a: str, b: str) -> float:
    """Similarity between a (possibly partial) quiz name and a full roster name."""
    ta, tb = _name_tokens(a), _name_tokens(b)
    if not ta or not tb:
        return 0.0
    matched = 0.0
    for q in ta:
        best = max((SequenceMatcher(None, q, r).ratio() for r in tb), default=0)
        if best >= 0.86:
            matched += 1
        elif best >= 0.72:
            matched += 0.5
    containment = matched / len(ta)
    first_bonus = 0.18 if SequenceMatcher(None, ta[0], tb[0]).ratio() >= 0.86 else 0.0
    seq = SequenceMatcher(None, ' '.join(ta), ' '.join(tb)).ratio()
    return round(min(1.0, containment * 0.62 + seq * 0.2 + first_bonus), 3)


async def _get_gradebook(gid: str, t):
    gb = await db.gradebooks.find_one({"id": gid, "owner_id": t["teacher_id"]}, {"_id": 0})
    if not gb:
        raise HTTPException(404, "سجل الدرجات غير موجود")
    return gb


# ---------- CRUD ----------

@api_router.get("/gradebooks")
async def list_gradebooks(t=Depends(get_teacher)):
    gbs = await db.gradebooks.find({"owner_id": t["teacher_id"], "year": teacher_year(t)}, {"_id": 0, "scores": 0}).to_list(200)
    for g in gbs:
        g["student_count"] = len(g.get("students", []))
        g.pop("students", None)
    return sorted(gbs, key=lambda x: (x.get("grade", ""), x.get("section", "")))


@api_router.post("/gradebooks")
async def create_gradebook(data: GradebookCreate, t=Depends(get_teacher)):
    if data.template not in GB_TEMPLATES:
        raise HTTPException(400, "نموذج السجل غير صحيح")
    existing = await db.gradebooks.find_one({"owner_id": t["teacher_id"], "year": teacher_year(t), "grade": data.grade.strip(), "section": data.section.strip()})
    if existing:
        raise HTTPException(400, "يوجد سجل بنفس الصف والشعبة مسبقاً")
    students = [{"id": str(uuid.uuid4()), "name": n.strip()} for n in data.students if n.strip()]
    gb = {
        "id": str(uuid.uuid4()),
        "owner_id": t["teacher_id"],
        "year": teacher_year(t),
        "template": data.template,
        "grade": data.grade.strip(),
        "section": data.section.strip(),
        "students": students,
        "scores": {"1": {}, "2": {}},
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    await db.gradebooks.insert_one(gb)
    gb.pop("_id", None)
    return gb


@api_router.get("/gradebooks/{gid}")
async def get_gradebook(gid: str, t=Depends(get_teacher)):
    return await _get_gradebook(gid, t)


@api_router.delete("/gradebooks/{gid}")
async def delete_gradebook(gid: str, t=Depends(get_teacher)):
    r = await db.gradebooks.delete_one({"id": gid, "owner_id": t["teacher_id"]})
    if r.deleted_count == 0:
        raise HTTPException(404, "سجل الدرجات غير موجود")
    return {"message": "تم الحذف"}


@api_router.post("/gradebooks/{gid}/students")
async def add_students(gid: str, data: GradebookStudentsAdd, t=Depends(get_teacher)):
    gb = await _get_gradebook(gid, t)
    existing = gb.get("students", [])
    existing_norm = {_norm_ar(s["name"]) for s in existing}
    added = []
    for n in data.names:
        n = n.strip()
        if n and _norm_ar(n) not in existing_norm:
            added.append({"id": str(uuid.uuid4()), "name": n})
            existing_norm.add(_norm_ar(n))
    if added:
        push = {"$each": added}
        if data.position is not None and 0 <= data.position <= len(existing):
            push["$position"] = data.position
        await db.gradebooks.update_one({"id": gid}, {"$push": {"students": push}, "$set": {"updated_at": now_iso()}})
    return {"added": len(added), "skipped": len(data.names) - len(added)}


@api_router.put("/gradebooks/{gid}/reorder")
async def reorder_students(gid: str, data: GradebookReorder, t=Depends(get_teacher)):
    gb = await _get_gradebook(gid, t)
    cur = {s["id"]: s for s in gb.get("students", [])}
    if set(data.student_ids) != set(cur.keys()) or len(data.student_ids) != len(cur):
        raise HTTPException(400, "قائمة الترتيب غير صالحة")
    new_students = [cur[i] for i in data.student_ids]
    await db.gradebooks.update_one({"id": gid}, {"$set": {"students": new_students, "updated_at": now_iso()}})
    return {"message": "تم تحديث الترتيب"}


@api_router.put("/gradebooks/{gid}/students/{sid}")
async def rename_student(gid: str, sid: str, data: GradebookStudentRename, t=Depends(get_teacher)):
    await _get_gradebook(gid, t)
    await db.gradebooks.update_one(
        {"id": gid, "students.id": sid},
        {"$set": {"students.$.name": data.name.strip(), "updated_at": now_iso()}}
    )
    return {"message": "تم التحديث"}


@api_router.delete("/gradebooks/{gid}/students/{sid}")
async def delete_student(gid: str, sid: str, t=Depends(get_teacher)):
    await _get_gradebook(gid, t)
    await db.gradebooks.update_one({"id": gid}, {
        "$pull": {"students": {"id": sid}},
        "$unset": {f"scores.1.{sid}": "", f"scores.2.{sid}": ""},
        "$set": {"updated_at": now_iso()}
    })
    return {"message": "تم الحذف"}


@api_router.put("/gradebooks/{gid}/scores")
async def update_scores(gid: str, data: GBScoresBulk, t=Depends(get_teacher)):
    gb = await _get_gradebook(gid, t)
    mx_map = gb_max_map(gb)
    valid_ids = {s["id"] for s in gb.get("students", [])}
    sets = {}
    unsets = {}
    for u in data.updates:
        if u.semester not in ("1", "2") or u.field not in mx_map or u.student_id not in valid_ids:
            continue
        key = f"scores.{u.semester}.{u.student_id}.{u.field}"
        if u.value is None:
            unsets[key] = ""
        else:
            v = max(0.0, min(float(u.value), float(mx_map[u.field])))
            sets[key] = v
    op = {}
    if sets:
        op["$set"] = {**sets, "updated_at": now_iso()}
    if unsets:
        op["$unset"] = unsets
        op.setdefault("$set", {"updated_at": now_iso()})
    if op:
        await db.gradebooks.update_one({"id": gid}, op)
    return {"saved": len(sets) + len(unsets)}


# ---------- Quiz → Gradebook smart sync ----------

@api_router.post("/gradebooks/{gid}/match-quiz")
async def match_quiz(gid: str, data: GBMatchReq, t=Depends(get_teacher)):
    gb = await _get_gradebook(gid, t)
    quiz = await db.quizzes.find_one({"id": data.quiz_id, "owner_id": t["teacher_id"]}, {"_id": 0, "id": 1, "title": 1})
    if not quiz:
        raise HTTPException(404, "الاختبار غير موجود")
    subs = await db.submissions.find(
        {"quiz_id": data.quiz_id, "submitted_at": {"$ne": None}},
        {"_id": 0, "id": 1, "student_name": 1, "grade": 1, "section": 1, "percentage": 1, "total_score": 1, "max_score": 1}
    ).to_list(500)
    students = gb.get("students", [])

    # compute all pair similarities, then greedy unique assignment (best first)
    pairs = []
    for si, sub in enumerate(subs):
        for st in students:
            score = name_similarity(sub["student_name"], st["name"])
            if score > 0:
                pairs.append((score, si, st["id"]))
    pairs.sort(key=lambda x: -x[0])

    assigned_subs = {}
    used_students = set()
    for score, si, stid in pairs:
        if score < 0.55:
            break
        if si in assigned_subs or stid in used_students:
            continue
        assigned_subs[si] = (stid, score)
        used_students.add(stid)

    proposals = []
    for si, sub in enumerate(subs):
        stid, conf = assigned_subs.get(si, (None, 0))
        proposals.append({
            **sub,
            "matched_student_id": stid,
            "confidence": conf
        })
    return {
        "quiz_title": quiz["title"],
        "students": students,
        "proposals": proposals
    }


@api_router.post("/gradebooks/{gid}/apply-quiz")
async def apply_quiz(gid: str, data: GBApplyReq, t=Depends(get_teacher)):
    gb = await _get_gradebook(gid, t)
    if data.semester not in ("1", "2"):
        raise HTTPException(400, "الفصل غير صحيح")
    mx_map = gb_max_map(gb)
    if data.column not in mx_map:
        raise HTTPException(400, "العمود غير صحيح أو غير متاح في نموذج هذا السجل")
    valid_ids = {s["id"] for s in gb.get("students", [])}
    sub_ids = [m.submission_id for m in data.mappings]
    subs = await db.submissions.find({"id": {"$in": sub_ids}, "quiz_id": data.quiz_id}, {"_id": 0, "id": 1, "percentage": 1}).to_list(500)
    pct_map = {s["id"]: s.get("percentage", 0) for s in subs}
    mx = float(mx_map[data.column])
    sets = {}
    for m in data.mappings:
        if m.student_id not in valid_ids or m.submission_id not in pct_map:
            continue
        score = round(pct_map[m.submission_id] * mx / 100 * 2) / 2  # nearest 0.5
        sets[f"scores.{data.semester}.{m.student_id}.{data.column}"] = score
    if sets:
        await db.gradebooks.update_one({"id": gid}, {"$set": {**sets, "updated_at": now_iso()}})
    return {"applied": len(sets)}


# ---------- Excel import / export ----------

def _parse_gb_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


@api_router.post("/gradebooks/import")
async def import_gradebook(data: GBImportReq, t=Depends(get_teacher)):
    try:
        content = base64.b64decode(data.data_base64)
        wb = _xl.load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "تعذر قراءة الملف — تأكد أنه ملف Excel صالح (xlsx)")

    s1 = next((wb[n] for n in wb.sheetnames if 'الفصل الأول' in n), None)
    if s1 is None:
        raise HTTPException(400, "الملف لا يطابق قالب سجل الدرجات (لا توجد ورقة الفصل الأول)")
    s2 = next((wb[n] for n in wb.sheetnames if 'الفصل الثاني' in n), None)

    # كشف نموذج الملف تلقائياً: في نموذج (7-10) العمود F = "الأنشطة العملية" والعمود I = "الاختبار القصير"
    f4 = str(s1['F4'].value or '')
    i4 = str(s1['I4'].value or '')
    tpl = "7-10" if ('الأنشطة' in f4 or 'الاختبار القصير' in i4) else "5-6"
    xl_cols = GB_XL_COLS_78 if tpl == "7-10" else GB_XL_COLS
    mx_map = GB_MAX_78 if tpl == "7-10" else GB_MAX

    m = _re.search(r'الصف\s*:\s*(.+)', str(s1['B2'].value or ''))
    grade = m.group(1).strip() if m else ''
    m = _re.search(r'الشعبة\s*:\s*(.+)', str(s1['B3'].value or ''))
    section = m.group(1).strip() if m else ''

    imported = []
    for r in range(6, min(s1.max_row + 1, 300)):
        name = s1.cell(r, 2).value
        if name is None or str(name).strip() == '':
            continue
        name = str(name).strip()
        if 'معلم المادة' in name or 'مشرف المادة' in name:
            break
        rec = {"name": name, "sem1": {}, "sem2": {}}
        for col, key in xl_cols:
            v = _parse_gb_number(s1.cell(r, col).value)
            if v is not None:
                rec["sem1"][key] = max(0.0, min(v, float(mx_map[key])))
            if s2 is not None:
                v2 = _parse_gb_number(s2.cell(r, col).value)
                if v2 is not None:
                    rec["sem2"][key] = max(0.0, min(v2, float(mx_map[key])))
        imported.append(rec)

    if not imported:
        raise HTTPException(400, "لم يتم العثور على أسماء طلاب في الملف")

    if data.gradebook_id:
        gb = await _get_gradebook(data.gradebook_id, t)
        if gb_template(gb) != tpl:
            raise HTTPException(400, f"الملف بنموذج الصفوف ({tpl}) بينما هذا السجل بنموذج ({gb_template(gb)}) — استورده في سجل مطابق أو من صفحة السجلات لإنشاء سجل جديد")
    else:
        if not grade:
            grade = 'الخامس' if tpl == "5-6" else 'الثامن'
        if not section:
            section = '1'
        gb = await db.gradebooks.find_one({"owner_id": t["teacher_id"], "year": teacher_year(t), "grade": grade, "section": section}, {"_id": 0})
        if gb and gb_template(gb) != tpl:
            raise HTTPException(400, f"يوجد سجل بنفس الصف والشعبة لكن بنموذج مختلف ({gb_template(gb)})")
        if not gb:
            gb = {
                "id": str(uuid.uuid4()), "owner_id": t["teacher_id"],
                "year": teacher_year(t),
                "template": tpl,
                "grade": grade, "section": section,
                "students": [], "scores": {"1": {}, "2": {}},
                "created_at": now_iso(), "updated_at": now_iso()
            }
            await db.gradebooks.insert_one(dict(gb))
            gb.pop("_id", None)

    norm_map = {_norm_ar(s["name"]): s["id"] for s in gb.get("students", [])}
    new_students = []
    sets = {}
    added = 0
    updated = 0
    for rec in imported:
        key = _norm_ar(rec["name"])
        sid = norm_map.get(key)
        if not sid:
            sid = str(uuid.uuid4())
            new_students.append({"id": sid, "name": rec["name"]})
            norm_map[key] = sid
            added += 1
        for sem, field_scores in (("1", rec["sem1"]), ("2", rec["sem2"])):
            for f, v in field_scores.items():
                sets[f"scores.{sem}.{sid}.{f}"] = v
                updated += 1
    op = {"$set": {**sets, "updated_at": now_iso()}}
    if new_students:
        op["$push"] = {"students": {"$each": new_students}}
    await db.gradebooks.update_one({"id": gb["id"]}, op)
    return {"gradebook_id": gb["id"], "students_added": added, "scores_imported": updated, "total_in_file": len(imported)}


def _gb_sem_sheet(wb, sheet_name, sem_label, gb, sem_key, teacher_name):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.rightToLeft = True
    thin = _Side(style='thin', color='9CA3AF')
    border = _Border(left=thin, right=thin, top=thin, bottom=thin)
    center = _Align(horizontal='center', vertical='center', wrap_text=True)
    right = _Align(horizontal='right', vertical='center')
    head_fill = _Fill('solid', fgColor='D9E1F2')
    max_fill = _Fill('solid', fgColor='FCE4D6')
    total_fill = _Fill('solid', fgColor='E2EFDA')
    bold = _Font(bold=True, size=11)

    ws.merge_cells('A1:O1')
    c = ws['A1']
    c.value = f'استمارة تقييم الطالب للصفوف(5-6) خلال الفصل الدراسي {sem_label}'
    c.font = _Font(bold=True, size=14)
    c.alignment = center
    ws['B2'] = f'الصف :  {gb["grade"]}'
    ws['B2'].font = bold
    ws['B3'] = f'الشعبة : {gb["section"]}'
    ws['B3'].font = bold

    ws.merge_cells('A4:A5')
    ws['A4'] = 'م'
    ws.merge_cells('B4:B5')
    ws['B4'] = 'إسم الطالب'
    ws.merge_cells('C4:D4')
    ws['C4'] = 'الحوار'
    ws['E4'] = 'المجموع'
    ws.merge_cells('F4:I4')
    ws['F4'] = 'الأسئلة القصيرة'
    ws['J4'] = 'المجموع'
    ws.merge_cells('K4:L4')
    ws['K4'] = 'الأنشطة العملية'
    ws['M4'] = 'المجموع'
    ws['N4'] = 'المشروع'
    ws['O4'] = 'الدرجة الكلية'
    maxes = {3: 10, 4: 10, 5: 20, 6: 5, 7: 5, 8: 5, 9: 5, 10: 20, 11: 20, 12: 20, 13: 40, 14: 20, 15: 100}
    for col, v in maxes.items():
        ws.cell(5, col, v)
    for r in (4, 5):
        for col in range(1, 16):
            cell = ws.cell(r, col)
            cell.alignment = center
            cell.border = border
            cell.font = bold
            cell.fill = head_fill if r == 4 else max_fill

    students = gb.get("students", [])
    scores = gb.get("scores", {}).get(sem_key, {})
    n_rows = max(40, len(students))
    for i in range(n_rows):
        r = 6 + i
        ws.cell(r, 1, i + 1)
        if i < len(students):
            st = students[i]
            ws.cell(r, 2, st["name"])
            sc = scores.get(st["id"], {})
            for col, key in GB_XL_COLS:
                v = sc.get(key)
                if v is not None:
                    ws.cell(r, col, v if v % 1 else int(v))
        ws.cell(r, 5).value = f'=SUM(C{r}:D{r})'
        ws.cell(r, 10).value = f'=SUM(F{r}+G{r}+H{r}+I{r})'
        ws.cell(r, 13).value = f'=IF(SUM(K{r}:L{r})=0,"",SUM(K{r}:L{r}))'
        ws.cell(r, 15).value = f'=SUM(E{r},J{r},M{r},N{r})'
        for col in range(1, 16):
            cell = ws.cell(r, col)
            cell.border = border
            cell.alignment = center if col != 2 else right
            if col in (5, 10, 13):
                cell.fill = total_fill
            if col == 15:
                cell.font = bold

    fr = 6 + n_rows
    ws.cell(fr, 2, f'معلم المادة/{teacher_name}')
    ws.merge_cells(f'C{fr}:K{fr}')
    ws.cell(fr, 3, 'يعتمد مشرف المادة/')
    ws.merge_cells(f'L{fr}:O{fr}')
    ws.cell(fr, 12, 'مدير المدرسة/')
    for col in (2, 3, 12):
        ws.cell(fr, col).font = bold

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 38
    for col in range(3, 16):
        ws.column_dimensions[_col_letter(col)].width = 9
    return ws


def _gb_sem_sheet78(wb, sheet_name, sem_label, gb, sem_key, teacher_name):
    """ورقة فصل دراسي بنموذج الصفوف (7-10): الحوار 20 + الأنشطة العملية 40 + الاختبار القصير 20 + المشروع 20"""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.rightToLeft = True
    thin = _Side(style='thin', color='9CA3AF')
    border = _Border(left=thin, right=thin, top=thin, bottom=thin)
    center = _Align(horizontal='center', vertical='center', wrap_text=True)
    right = _Align(horizontal='right', vertical='center')
    head_fill = _Fill('solid', fgColor='D9E1F2')
    max_fill = _Fill('solid', fgColor='FCE4D6')
    total_fill = _Fill('solid', fgColor='E2EFDA')
    bold = _Font(bold=True, size=11)

    ws.merge_cells('A1:K1')
    c = ws['A1']
    c.value = f'استمارة تقييم الطالب للصفوف (7-10) خلال الفصل الدراسي {sem_label}'
    c.font = _Font(bold=True, size=14)
    c.alignment = center
    ws['B2'] = f'الصف :  {gb["grade"]}'
    ws['B2'].font = bold
    ws['B3'] = f'الشعبة : {gb["section"]}'
    ws['B3'].font = bold

    ws.merge_cells('A4:A5')
    ws['A4'] = 'م'
    ws.merge_cells('B4:B5')
    ws['B4'] = 'إسم الطالب'
    ws.merge_cells('C4:D4')
    ws['C4'] = 'الحوار'
    ws['E4'] = 'المجموع'
    ws.merge_cells('F4:G4')
    ws['F4'] = 'الأنشطة العملية'
    ws['H4'] = 'المجموع'
    ws['I4'] = 'الاختبار القصير'
    ws['J4'] = 'المشروع'
    ws['K4'] = 'الدرجة الكلية'
    maxes = {3: 10, 4: 10, 5: 20, 6: 20, 7: 20, 8: 40, 9: 20, 10: 20, 11: 100}
    for col, v in maxes.items():
        ws.cell(5, col, v)
    for r in (4, 5):
        for col in range(1, 12):
            cell = ws.cell(r, col)
            cell.alignment = center
            cell.border = border
            cell.font = bold
            cell.fill = head_fill if r == 4 else max_fill

    students = gb.get("students", [])
    scores = gb.get("scores", {}).get(sem_key, {})
    n_rows = max(40, len(students))
    for i in range(n_rows):
        r = 6 + i
        ws.cell(r, 1, i + 1)
        if i < len(students):
            st = students[i]
            ws.cell(r, 2, st["name"])
            sc = scores.get(st["id"], {})
            for col, key in GB_XL_COLS_78:
                v = sc.get(key)
                if v is not None:
                    ws.cell(r, col, v if v % 1 else int(v))
        ws.cell(r, 5).value = f'=SUM(C{r}:D{r})'
        ws.cell(r, 8).value = f'=SUM(F{r}:G{r})'
        ws.cell(r, 11).value = f'=SUM(E{r},H{r},I{r},J{r})'
        for col in range(1, 12):
            cell = ws.cell(r, col)
            cell.border = border
            cell.alignment = center if col != 2 else right
            if col in (5, 8):
                cell.fill = total_fill
            if col == 11:
                cell.font = bold

    fr = 6 + n_rows
    ws.cell(fr, 2, f'معلم المادة/{teacher_name}')
    ws.merge_cells(f'C{fr}:F{fr}')
    ws.cell(fr, 3, 'يعتمد مشرف المادة/')
    ws.merge_cells(f'G{fr}:K{fr}')
    ws.cell(fr, 7, 'مدير المدرسة/')
    for col in (2, 3, 7):
        ws.cell(fr, col).font = bold

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 38
    for col in range(3, 12):
        ws.column_dimensions[_col_letter(col)].width = 11
    return ws


def _gb_annual_sheet(wb, gb):
    ws = wb.create_sheet('سجل سنوي')
    ws.sheet_view.rightToLeft = True
    thin = _Side(style='thin', color='9CA3AF')
    border = _Border(left=thin, right=thin, top=thin, bottom=thin)
    center = _Align(horizontal='center', vertical='center', wrap_text=True)
    right = _Align(horizontal='right', vertical='center')
    head_fill = _Fill('solid', fgColor='D9E1F2')
    bold = _Font(bold=True, size=11)

    year = datetime.now(timezone.utc).year
    is78 = gb_template(gb) == "7-10"
    grades_label = '(7 -10)' if is78 else '(5 - 6)'
    ws.merge_cells('A1:K2')
    ws['A1'] = f'استمارة تقييم طلبة الحلقة الثانية للصفوف {grades_label} لمادة تقنية المعلومات للعام الدراسي {year} / {year + 1}م'
    ws['A1'].font = _Font(bold=True, size=13)
    ws['A1'].alignment = center
    ws.merge_cells('A3:K3')
    ws['A3'] = 'السجــــــــــل السنــــــــــــــــــوي'
    ws['A3'].font = _Font(bold=True, size=13)
    ws['A3'].alignment = center

    ws.merge_cells('A4:A7')
    ws['A4'] = 'م'
    ws.merge_cells('B4:B7')
    ws['B4'] = 'اسم الطالب'
    ws.merge_cells('C4:F4')
    ws['C4'] = 'الفصل الدراسي الأول '
    ws.merge_cells('G4:I4')
    ws['G4'] = 'الفصل الدراسي الثاني'
    ws.merge_cells('J4:J6')
    ws['J4'] = 'درجة نهاية العام'
    ws.merge_cells('K4:K7')
    ws['K4'] = 'المستوى'
    ws.merge_cells('C5:C6')
    ws['C5'] = 'درجة منتصف الفصل'
    ws.merge_cells('D5:D7')
    ws['D5'] = 'العبارة الوصفية'
    ws.merge_cells('E5:E6')
    ws['E5'] = 'درجة نهاية الفصل'
    ws.merge_cells('F5:F7')
    ws['F5'] = 'المستوى'
    ws.merge_cells('G5:G6')
    ws['G5'] = 'درجة منتصف الفصل'
    ws.merge_cells('H5:H7')
    ws['H5'] = 'العبارة الوصفية'
    ws.merge_cells('I5:I6')
    ws['I5'] = 'درجة نهاية الفصل'
    for col, v in [(3, 100), (5, 100), (7, 100), (9, 100), (10, 100)]:
        ws.cell(7, col, v)
    for r in range(4, 8):
        for col in range(1, 12):
            cell = ws.cell(r, col)
            cell.alignment = center
            cell.border = border
            cell.font = bold
            cell.fill = head_fill

    s1, s2 = GB_SEM1, GB_SEM2
    # صيغ منتصف الفصل ونهاية الفصل حسب النموذج:
    # 5-6: منتصف = (حوار1 + قصيرة1 + قصيرة2 + عملي1) ×100/40، النهاية = العمود O
    # 7-10: منتصف = (حوار1 + عملي1 + الاختبار القصير) ×100/50، النهاية = العمود K
    if is78:
        def mid_formula(sheet, sr):
            return f"=IF('{sheet}'!C{sr}=\"\",\"\",SUM('{sheet}'!C{sr},'{sheet}'!F{sr},'{sheet}'!I{sr})*100/50)"
        total_col = 'K'
    else:
        def mid_formula(sheet, sr):
            return f"=IF('{sheet}'!C{sr}=\"\",\"\",SUM('{sheet}'!C{sr},'{sheet}'!F{sr},'{sheet}'!G{sr},'{sheet}'!K{sr})*100/40)"
        total_col = 'O'
    n_rows = max(40, len(gb.get("students", [])))
    for i in range(n_rows):
        r = 8 + i
        sr = 6 + i
        ws.cell(r, 1, i + 1)
        ws.cell(r, 2).value = f"='{s1}'!B{sr}"
        ws.cell(r, 3).value = mid_formula(s1, sr)
        ws.cell(r, 4).value = f'=IF(C{r}="","",IF(C{r}>=95,"1",IF(C{r}>=90,"2",IF(C{r}>=80,"3",IF(C{r}>=70,"4",IF(C{r}>=60,"5",IF(C{r}>=50,"6",IF(C{r}<50,"7"))))))))'
        ws.cell(r, 5).value = f"='{s1}'!{total_col}{sr}"
        ws.cell(r, 6).value = f'=IF(E{r}="","",IF(E{r}>=90,"أ",IF(E{r}>=80,"ب",IF(E{r}>=65,"ج",IF(E{r}>=50,"د",IF(E{r}<50,"هـ"))))))'
        ws.cell(r, 7).value = mid_formula(s2, sr)
        ws.cell(r, 8).value = f'=IF(G{r}="","",IF(G{r}>=95,"1",IF(G{r}>=90,"2",IF(G{r}>=80,"3",IF(G{r}>=70,"4",IF(G{r}>=60,"5",IF(G{r}>=50,"6",IF(G{r}<50,"7"))))))))'
        ws.cell(r, 9).value = f"='{s2}'!{total_col}{sr}"
        ws.cell(r, 10).value = f'=IF(I{r}="","",(AVERAGE(E{r},I{r})))'
        ws.cell(r, 11).value = f'=IF(J{r}="","",IF(J{r}>=90,"أ",IF(J{r}>=80,"ب",IF(J{r}>=65,"ج",IF(J{r}>=50,"د",IF(J{r}<50,"هـ"))))))'
        for col in range(1, 12):
            cell = ws.cell(r, col)
            cell.border = border
            cell.alignment = center if col != 2 else right

    fr = 8 + n_rows
    ws.merge_cells(f'A{fr}:C{fr}')
    ws.cell(fr, 1, '    معلم المادة/')
    ws.merge_cells(f'D{fr}:E{fr}')
    ws.cell(fr, 4, 'مشرف المادة/ ')
    ws.merge_cells(f'H{fr}:J{fr}')
    ws.cell(fr, 8, 'مدير المدرسة/')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 38
    for col in range(3, 12):
        ws.column_dimensions[_col_letter(col)].width = 11
    return ws


def _gb_project_sheet(wb, gb):
    ws = wb.create_sheet('استمارة المشروع المقترح الأول ')
    ws.sheet_view.rightToLeft = True
    thin = _Side(style='thin', color='9CA3AF')
    border = _Border(left=thin, right=thin, top=thin, bottom=thin)
    center = _Align(horizontal='center', vertical='center', wrap_text=True)
    right = _Align(horizontal='right', vertical='center')
    head_fill = _Fill('solid', fgColor='D9E1F2')
    bold = _Font(bold=True, size=11)

    ws.merge_cells('A1:G2')
    ws['A1'] = f'استمارة تقييم المشروع في مادة تقنية المعلومات للصف ( {gb["grade"]} / {gb["section"]} )'
    ws['A1'].font = _Font(bold=True, size=13)
    ws['A1'].alignment = center
    ws.merge_cells('A3:A5')
    ws['A3'] = 'م'
    ws.merge_cells('B3:B5')
    ws['B3'] = 'اسم الطالب'
    ws.merge_cells('C3:E3')
    ws['C3'] = 'تطبيق العمليات'
    ws['F3'] = 'حل المشكلات'
    ws.merge_cells('G3:G4')
    ws['G3'] = 'المجموع'
    ws['C4'] = 'التخطيط'
    ws['D4'] = 'التنفيذ العملي'
    ws['E4'] = 'الاتصال وعمل الفريق'
    ws['F4'] = 'التحليل والاستنتاج'
    for col, v in [(3, 5), (4, 7), (5, 3), (6, 5)]:
        ws.cell(5, col, v)
    ws.cell(5, 7).value = '=SUM(C5:F5)'
    for r in range(3, 6):
        for col in range(1, 8):
            cell = ws.cell(r, col)
            cell.alignment = center
            cell.border = border
            cell.font = bold
            cell.fill = head_fill

    n_rows = max(40, len(gb.get("students", [])))
    for i in range(n_rows):
        r = 6 + i
        ws.cell(r, 1, i + 1)
        ws.cell(r, 2).value = f"='{GB_SEM1}'!B{6 + i}"
        for col in range(1, 8):
            cell = ws.cell(r, col)
            cell.border = border
            cell.alignment = center if col != 2 else right
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 38
    for col in range(3, 8):
        ws.column_dimensions[_col_letter(col)].width = 14
    return ws


@api_router.get("/gradebooks/{gid}/export")
async def export_gradebook(gid: str, t=Depends(get_teacher)):
    gb = await _get_gradebook(gid, t)
    teacher_name = t["teacher"].get("teacher_name", "")
    wb = _xl.Workbook()
    wb.remove(wb.active)
    sem_sheet = _gb_sem_sheet78 if gb_template(gb) == "7-10" else _gb_sem_sheet
    sem_sheet(wb, GB_SEM1, 'الأول', gb, "1", teacher_name)
    sem_sheet(wb, GB_SEM2, 'الثاني', gb, "2", teacher_name)
    _gb_annual_sheet(wb, gb)
    _gb_project_sheet(wb, gb)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="gradebook.xlsx"'}
    )


# =================== ADMIN: TEACHER MANAGEMENT ===================

class TeacherCreateReq(BaseModel):
    username: str
    password: str
    teacher_name: str
    school_name: Optional[str] = None


class TeacherUpdateReq(BaseModel):
    teacher_name: Optional[str] = None
    school_name: Optional[str] = None
    new_password: Optional[str] = None
    is_active: Optional[bool] = None


@api_router.get("/admin/teachers")
async def admin_list_teachers(_=Depends(require_admin)):
    teachers = await db.teachers.find({}, {"_id": 0, "password_hash": 0}).to_list(200)
    for tc in teachers:
        tc["quiz_count"] = await db.quizzes.count_documents({"owner_id": tc["id"]})
        tc["project_count"] = await db.projects.count_documents({"owner_id": tc["id"]})
    return sorted(teachers, key=lambda x: (x.get("role") != "admin", x.get("created_at", "")))


@api_router.post("/admin/teachers")
async def admin_create_teacher(req: TeacherCreateReq, _=Depends(require_admin)):
    username = req.username.strip().lower()
    if len(username) < 3:
        raise HTTPException(400, "اسم المستخدم يجب أن يكون 3 أحرف على الأقل")
    if not username.replace("_", "").replace(".", "").isalnum() or not username.isascii():
        raise HTTPException(400, "اسم المستخدم يجب أن يحتوي على أحرف إنجليزية وأرقام فقط")
    if len(req.password) < 6:
        raise HTTPException(400, "كلمة المرور يجب أن تكون 6 أحرف على الأقل")
    if await db.teachers.find_one({"username": username}):
        raise HTTPException(400, "اسم المستخدم موجود مسبقاً")
    teacher = {
        "id": str(uuid.uuid4()),
        "username": username,
        "password_hash": hash_password(req.password),
        "teacher_name": req.teacher_name.strip() or "معلم جديد",
        "school_name": (req.school_name or "").strip() or "مدرسة الخيرات للتعليم الأساسي",
        "role": "teacher",
        "is_active": True,
        "created_at": now_iso()
    }
    await db.teachers.insert_one(teacher)
    teacher.pop("_id", None)
    teacher.pop("password_hash", None)
    teacher["quiz_count"] = 0
    teacher["project_count"] = 0
    return teacher


@api_router.put("/admin/teachers/{tid}")
async def admin_update_teacher(tid: str, req: TeacherUpdateReq, _=Depends(require_admin)):
    doc = await db.teachers.find_one({"id": tid})
    if not doc:
        raise HTTPException(404, "المعلم غير موجود")
    upd = {}
    if req.teacher_name is not None and req.teacher_name.strip():
        upd["teacher_name"] = req.teacher_name.strip()
    if req.school_name is not None:
        upd["school_name"] = req.school_name.strip() or "مدرسة الخيرات للتعليم الأساسي"
    if req.new_password:
        if len(req.new_password) < 6:
            raise HTTPException(400, "كلمة المرور يجب أن تكون 6 أحرف على الأقل")
        upd["password_hash"] = hash_password(req.new_password)
    if req.is_active is not None:
        if doc.get("role") == "admin":
            raise HTTPException(400, "لا يمكن تعطيل حساب المدير")
        upd["is_active"] = req.is_active
    if upd:
        await db.teachers.update_one({"id": tid}, {"$set": upd})
    return {"message": "تم التحديث"}


@api_router.delete("/admin/teachers/{tid}")
async def admin_delete_teacher(tid: str, _=Depends(require_admin)):
    doc = await db.teachers.find_one({"id": tid})
    if not doc:
        raise HTTPException(404, "المعلم غير موجود")
    if doc.get("role") == "admin":
        raise HTTPException(400, "لا يمكن حذف حساب المدير")
    quiz_ids = [q["id"] for q in await db.quizzes.find({"owner_id": tid}, {"id": 1}).to_list(2000)]
    if quiz_ids:
        await db.submissions.delete_many({"quiz_id": {"$in": quiz_ids}})
        await db.quizzes.delete_many({"owner_id": tid})
    project_ids = [p["id"] for p in await db.projects.find({"owner_id": tid}, {"id": 1}).to_list(2000)]
    if project_ids:
        sub_ids = [s["id"] for s in await db.project_submissions.find({"project_id": {"$in": project_ids}}, {"id": 1}).to_list(5000)]
        if sub_ids:
            await db.project_files.delete_many({"submission_id": {"$in": sub_ids}})
        await db.project_submissions.delete_many({"project_id": {"$in": project_ids}})
        await db.projects.delete_many({"owner_id": tid})
    await db.question_bank.delete_many({"owner_id": tid})
    await db.teachers.delete_one({"id": tid})
    return {"message": "تم حذف المعلم وجميع بياناته"}


# =================== بطاقات التقييم السريع (Rubrics) ===================

class RubricCriterion(BaseModel):
    id: Optional[str] = None
    name: str
    max: float

class RubricCreate(BaseModel):
    title: str
    semester: str = "1"
    column: str
    criteria: List[RubricCriterion]
    images: Optional[List[str]] = None  # قائمة data URLs للصور المرفقة بالبطاقة (اختياري)

class RubricEvalSave(BaseModel):
    gradebook_id: str
    student_id: str
    scores: dict


def _clean_criteria(criteria):
    out = []
    for c in criteria:
        name = c.name.strip()
        if not name or c.max <= 0:
            continue
        out.append({"id": c.id or str(uuid.uuid4()), "name": name, "max": float(c.max)})
    if not out:
        raise HTTPException(400, "أضف معياراً واحداً على الأقل")
    return out


def _validate_rubric_meta(data: RubricCreate):
    if data.column not in GB_MAX:
        raise HTTPException(400, "عمود السجل غير صحيح")
    if data.semester not in ("1", "2"):
        raise HTTPException(400, "الفصل غير صحيح")
    if not data.title.strip():
        raise HTTPException(400, "اكتب عنوان بطاقة التقييم")


async def _get_rubric(rid: str, t):
    r = await db.rubrics.find_one({"id": rid, "owner_id": t["teacher_id"]}, {"_id": 0})
    if not r:
        raise HTTPException(404, "بطاقة التقييم غير موجودة")
    return r


@api_router.get("/rubrics")
async def list_rubrics(t=Depends(get_teacher)):
    rubrics = await db.rubrics.find({"owner_id": t["teacher_id"], "year": teacher_year(t)}, {"_id": 0}).to_list(300)
    counts = await db.rubric_evaluations.aggregate([
        {"$match": {"owner_id": t["teacher_id"]}},
        {"$group": {"_id": "$rubric_id", "n": {"$sum": 1}}}
    ]).to_list(500)
    cmap = {c["_id"]: c["n"] for c in counts}
    for r in rubrics:
        r["evaluation_count"] = cmap.get(r["id"], 0)
    return sorted(rubrics, key=lambda x: x.get("created_at", ""), reverse=True)


@api_router.post("/rubrics")
async def create_rubric(data: RubricCreate, t=Depends(get_teacher)):
    _validate_rubric_meta(data)
    criteria = _clean_criteria(data.criteria)
    images = [img for img in (data.images or []) if isinstance(img, str) and img.startswith("data:")][:8]
    rubric = {
        "id": str(uuid.uuid4()),
        "owner_id": t["teacher_id"],
        "year": teacher_year(t),
        "title": data.title.strip(),
        "semester": data.semester,
        "column": data.column,
        "criteria": criteria,
        "images": images,
        "total_max": round(sum(c["max"] for c in criteria), 2),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.rubrics.insert_one(dict(rubric))
    return rubric


@api_router.get("/rubrics/{rid}")
async def get_rubric(rid: str, t=Depends(get_teacher)):
    return await _get_rubric(rid, t)


@api_router.put("/rubrics/{rid}")
async def update_rubric(rid: str, data: RubricCreate, t=Depends(get_teacher)):
    await _get_rubric(rid, t)
    _validate_rubric_meta(data)
    criteria = _clean_criteria(data.criteria)
    images = [img for img in (data.images or []) if isinstance(img, str) and img.startswith("data:")][:8]
    await db.rubrics.update_one({"id": rid}, {"$set": {
        "title": data.title.strip(), "semester": data.semester, "column": data.column,
        "criteria": criteria, "images": images,
        "total_max": round(sum(c["max"] for c in criteria), 2),
        "updated_at": now_iso(),
    }})
    return await _get_rubric(rid, t)


@api_router.delete("/rubrics/{rid}")
async def delete_rubric(rid: str, t=Depends(get_teacher)):
    await _get_rubric(rid, t)
    await db.rubric_evaluations.delete_many({"rubric_id": rid})
    await db.rubrics.delete_one({"id": rid})
    return {"message": "تم الحذف"}


@api_router.get("/rubrics/{rid}/evaluations")
async def list_rubric_evaluations(rid: str, gradebook_id: str, t=Depends(get_teacher)):
    await _get_rubric(rid, t)
    return await db.rubric_evaluations.find(
        {"rubric_id": rid, "gradebook_id": gradebook_id, "owner_id": t["teacher_id"]}, {"_id": 0}
    ).to_list(500)


@api_router.put("/rubrics/{rid}/evaluations")
async def save_rubric_evaluation(rid: str, data: RubricEvalSave, t=Depends(get_teacher)):
    rubric = await _get_rubric(rid, t)
    gb = await _get_gradebook(data.gradebook_id, t)
    if data.student_id not in {s["id"] for s in gb.get("students", [])}:
        raise HTTPException(404, "الطالب غير موجود في السجل")
    crit = {c["id"]: c["max"] for c in rubric["criteria"]}
    scores = {}
    for cid, v in (data.scores or {}).items():
        if cid not in crit or v is None:
            continue
        scores[cid] = max(0.0, min(float(v), crit[cid]))
    total = round(sum(scores.values()), 2)
    # تحويل المجموع إلى درجة عمود السجل (تحجيم تلقائي إذا اختلف الحد الأقصى) — حسب نموذج السجل
    mx_map = gb_max_map(gb)
    if rubric["column"] not in mx_map:
        raise HTTPException(400, "عمود البطاقة غير متاح في نموذج هذا السجل")
    col_max = float(mx_map[rubric["column"]])
    tot_max = float(rubric.get("total_max") or 0) or col_max
    gb_score = round((total / tot_max) * col_max * 2) / 2
    gb_score = max(0.0, min(gb_score, col_max))
    await db.rubric_evaluations.update_one(
        {"rubric_id": rid, "gradebook_id": data.gradebook_id, "student_id": data.student_id},
        {"$set": {
            "owner_id": t["teacher_id"], "scores": scores, "total": total,
            "gb_score": gb_score,
            "updated_at": now_iso(),
        }, "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now_iso()}},
        upsert=True
    )
    # النقل التلقائي إلى سجل الدرجات
    await db.gradebooks.update_one(
        {"id": data.gradebook_id},
        {"$set": {f"scores.{rubric['semester']}.{data.student_id}.{rubric['column']}": gb_score, "updated_at": now_iso()}}
    )
    return {"total": total, "gb_score": gb_score, "column": rubric["column"], "semester": rubric["semester"]}


# =================== GRADE RELEASES (إرسال نتائج تقييم نشاط/مشروع للطلاب مع المعايير) ===================

RELEASE_COLUMNS = {"p1": "النشاط الأول", "p2": "النشاط الثاني", "proj": "المشروع"}


def _col_label(gb, column):
    fields = GB_FIELDS_78 if gb_template(gb) == "7-10" else GB_FIELDS
    for k, label, _ in fields:
        if k == column:
            return RELEASE_COLUMNS.get(column, label)
    return RELEASE_COLUMNS.get(column, column)


class GradeReleaseCreate(BaseModel):
    rubric_id: str
    gradebook_id: str


class StudentGradesCheck(BaseModel):
    student_name: str
    grade: str
    section: str


@api_router.get("/grade-releases")
async def list_grade_releases(t=Depends(get_teacher)):
    rels = await db.grade_releases.find({"owner_id": t["teacher_id"]}, {"_id": 0}).to_list(100)
    return sorted(rels, key=lambda x: x.get("created_at", ""), reverse=True)


@api_router.post("/grade-releases")
async def create_grade_release(data: GradeReleaseCreate, t=Depends(get_teacher)):
    rubric = await _get_rubric(data.rubric_id, t)
    gb = await _get_gradebook(data.gradebook_id, t)
    if rubric["column"] not in gb_max_map(gb):
        raise HTTPException(400, "عمود البطاقة غير متاح في نموذج هذا السجل")
    rel = {
        "owner_id": t["teacher_id"],
        "rubric_id": rubric["id"],
        "rubric_title": rubric["title"],
        "column": rubric["column"],
        "column_label": _col_label(gb, rubric["column"]),
        "semester": rubric["semester"],
        "gradebook_id": gb["id"],
        "grade": gb["grade"],
        "section": str(gb["section"]),
        "created_at": now_iso(),
    }
    await db.grade_releases.update_one(
        {"rubric_id": rubric["id"], "gradebook_id": gb["id"]},
        {"$set": rel, "$setOnInsert": {"id": str(uuid.uuid4())}},
        upsert=True
    )
    doc = await db.grade_releases.find_one(
        {"rubric_id": rubric["id"], "gradebook_id": gb["id"]}, {"_id": 0}
    )
    return doc


@api_router.delete("/grade-releases/{rid}")
async def delete_grade_release(rid: str, t=Depends(get_teacher)):
    r = await db.grade_releases.delete_one({"id": rid, "owner_id": t["teacher_id"]})
    if r.deleted_count == 0:
        raise HTTPException(404, "غير موجود")
    return {"message": "تم إيقاف عرض الدرجات"}


@api_router.post("/student/check-grades")
async def student_check_grades(data: StudentGradesCheck):
    """صفحة الطالب العامة: يفحص هل أرسل المعلم نتائج تقييم لصفه/شعبته — يعرض المعايير ودرجة كل معيار"""
    name = data.student_name.strip()
    if not name:
        raise HTTPException(400, "اكتب اسمك")
    rels = await db.grade_releases.find(
        {"grade": data.grade.strip(), "section": str(data.section).strip()}, {"_id": 0}
    ).to_list(50)
    if not rels:
        return {"waiting": True}
    results = []
    matched_name = None
    gb_cache = {}
    for rel in rels:
        gb = gb_cache.get(rel["gradebook_id"])
        if gb is None:
            gb = await db.gradebooks.find_one({"id": rel["gradebook_id"]}, {"_id": 0})
            gb_cache[rel["gradebook_id"]] = gb or False
        if not gb:
            continue
        best, best_score = None, 0
        for st in gb.get("students", []):
            s = name_similarity(name, st["name"])
            if s > best_score:
                best, best_score = st, s
        if not best or best_score < 0.55:
            continue
        matched_name = best["name"]
        rubric = await db.rubrics.find_one({"id": rel["rubric_id"]}, {"_id": 0})
        if not rubric:
            continue
        ev = await db.rubric_evaluations.find_one(
            {"rubric_id": rel["rubric_id"], "gradebook_id": rel["gradebook_id"], "student_id": best["id"]},
            {"_id": 0}
        )
        ev_scores = (ev or {}).get("scores", {})
        criteria = [
            {"name": c["name"], "max": c["max"], "score": ev_scores.get(c["id"])}
            for c in rubric.get("criteria", [])
        ]
        results.append({
            "title": rel.get("rubric_title", rubric["title"]),
            "column": rel["column"],
            "column_label": rel.get("column_label", RELEASE_COLUMNS.get(rel["column"], rel["column"])),
            "semester": rel["semester"],
            "evaluated": ev is not None,
            "criteria": criteria,
            "total": (ev or {}).get("total"),
            "total_max": rubric.get("total_max"),
            "gb_score": (ev or {}).get("gb_score"),
            "gb_max": gb_max_map(gb).get(rel["column"]),
        })
    if not results:
        return {"waiting": False, "found": False,
                "message": "لم يتم العثور على اسمك في سجل الصف — تأكد من كتابة اسمك كاملاً كما هو مسجل لدى المعلم"}
    return {"waiting": False, "found": True, "student_name": matched_name, "grades": results}


# =================== GRADE SHARING SESSIONS (رابط مشاركة + مطابقة يدوية + إرسال نتائج) ===================

class GradeSessionCreate(BaseModel):
    rubric_id: str
    gradebook_id: str


class SessionJoinRequest(BaseModel):
    name: str
    grade: str
    section: str


class SessionMatchUpdate(BaseModel):
    matched_student_id: Optional[str] = None
    confirmed: Optional[bool] = None
    ignored: Optional[bool] = None


def _gen_session_code() -> str:
    """رمز جلسة قصير من 6 أحرف (بدون أحرف مربكة)"""
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "".join(random.choices(alphabet, k=6))


def _auto_match_participant(name: str, gb_students: list) -> dict:
    """يرجع أفضل مطابقة من السجل لاسم الطالب الداخل"""
    best, best_score = None, 0.0
    for st in gb_students:
        s = name_similarity(name, st["name"])
        if s > best_score:
            best, best_score = st, s
    if not best or best_score < 0.45:
        return {"matched_student_id": None, "matched_student_name": None, "match_confidence": round(best_score, 2)}
    return {
        "matched_student_id": best["id"],
        "matched_student_name": best["name"],
        "match_confidence": round(best_score, 2),
    }


def _session_public_view(sess: dict) -> dict:
    """نسخة عامة من الجلسة (للطالب) بدون معلومات حساسة"""
    return {
        "code": sess["code"],
        "rubric_title": sess.get("rubric_title"),
        "column_label": sess.get("column_label"),
        "grade": sess.get("grade"),
        "section": sess.get("section"),
        "status": sess.get("status", "open"),
    }


@api_router.post("/grade-sessions")
async def create_grade_session(data: GradeSessionCreate, t=Depends(get_teacher)):
    rubric = await _get_rubric(data.rubric_id, t)
    gb = await _get_gradebook(data.gradebook_id, t)
    if rubric["column"] not in gb_max_map(gb):
        raise HTTPException(400, "عمود البطاقة غير متاح في نموذج هذا السجل")
    # رمز فريد
    code = _gen_session_code()
    while await db.grade_sessions.find_one({"code": code}):
        code = _gen_session_code()
    sess = {
        "id": str(uuid.uuid4()),
        "code": code,
        "owner_id": t["teacher_id"],
        "rubric_id": rubric["id"],
        "rubric_title": rubric["title"],
        "column": rubric["column"],
        "column_label": _col_label(gb, rubric["column"]),
        "semester": rubric["semester"],
        "gradebook_id": gb["id"],
        "grade": gb["grade"],
        "section": str(gb["section"]),
        "status": "open",  # open | released | closed
        "participants": [],
        "released_at": None,
        "created_at": now_iso(),
    }
    await db.grade_sessions.insert_one(sess)
    sess.pop("_id", None)
    return sess


@api_router.get("/grade-sessions")
async def list_grade_sessions(t=Depends(get_teacher)):
    sessions = await db.grade_sessions.find(
        {"owner_id": t["teacher_id"]}, {"_id": 0}
    ).to_list(200)
    for s in sessions:
        s["participants_count"] = len(s.get("participants", []))
    return sorted(sessions, key=lambda x: x.get("created_at", ""), reverse=True)


@api_router.get("/grade-sessions/{sid}")
async def get_grade_session(sid: str, t=Depends(get_teacher)):
    sess = await db.grade_sessions.find_one({"id": sid, "owner_id": t["teacher_id"]}, {"_id": 0})
    if not sess:
        raise HTTPException(404, "الجلسة غير موجودة")
    # إضافة قائمة طلاب السجل لتمكين المطابقة اليدوية
    gb = await db.gradebooks.find_one({"id": sess["gradebook_id"]}, {"_id": 0, "students": 1})
    sess["roster"] = (gb or {}).get("students", [])
    # إضافة تفاصيل التقييم (الدرجات الموجودة) لكل طالب مطابق
    evals = await db.rubric_evaluations.find(
        {"rubric_id": sess["rubric_id"], "gradebook_id": sess["gradebook_id"]}, {"_id": 0}
    ).to_list(500)
    eval_map = {ev["student_id"]: ev for ev in evals}
    for p in sess.get("participants", []):
        sid_match = p.get("matched_student_id")
        if sid_match and sid_match in eval_map:
            ev = eval_map[sid_match]
            p["has_evaluation"] = True
            p["preview_total"] = ev.get("total")
            p["preview_gb_score"] = ev.get("gb_score")
        else:
            p["has_evaluation"] = False
    return sess


@api_router.delete("/grade-sessions/{sid}")
async def delete_grade_session(sid: str, t=Depends(get_teacher)):
    r = await db.grade_sessions.delete_one({"id": sid, "owner_id": t["teacher_id"]})
    if r.deleted_count == 0:
        raise HTTPException(404, "الجلسة غير موجودة")
    return {"message": "تم حذف الجلسة"}


@api_router.put("/grade-sessions/{sid}/participants/{pid}")
async def update_participant_match(sid: str, pid: str, data: SessionMatchUpdate, t=Depends(get_teacher)):
    sess = await db.grade_sessions.find_one({"id": sid, "owner_id": t["teacher_id"]})
    if not sess:
        raise HTTPException(404, "الجلسة غير موجودة")
    participants = sess.get("participants", [])
    target = next((p for p in participants if p["id"] == pid), None)
    if not target:
        raise HTTPException(404, "المشارك غير موجود")
    if data.matched_student_id is not None:
        if data.matched_student_id == "":
            target["matched_student_id"] = None
            target["matched_student_name"] = None
        else:
            gb = await db.gradebooks.find_one({"id": sess["gradebook_id"]}, {"_id": 0, "students": 1})
            st = next((s for s in (gb or {}).get("students", []) if s["id"] == data.matched_student_id), None)
            if not st:
                raise HTTPException(400, "الطالب غير موجود في السجل")
            target["matched_student_id"] = st["id"]
            target["matched_student_name"] = st["name"]
    if data.confirmed is not None:
        target["confirmed"] = bool(data.confirmed)
        if data.confirmed:
            target["ignored"] = False
    if data.ignored is not None:
        target["ignored"] = bool(data.ignored)
        if data.ignored:
            target["confirmed"] = False
    await db.grade_sessions.update_one({"id": sid}, {"$set": {"participants": participants}})
    return target


@api_router.delete("/grade-sessions/{sid}/participants/{pid}")
async def remove_participant(sid: str, pid: str, t=Depends(get_teacher)):
    r = await db.grade_sessions.update_one(
        {"id": sid, "owner_id": t["teacher_id"]},
        {"$pull": {"participants": {"id": pid}}}
    )
    if r.modified_count == 0:
        raise HTTPException(404, "المشارك غير موجود")
    return {"message": "تمت الإزالة"}


@api_router.post("/grade-sessions/{sid}/release")
async def release_session(sid: str, t=Depends(get_teacher)):
    sess = await db.grade_sessions.find_one({"id": sid, "owner_id": t["teacher_id"]})
    if not sess:
        raise HTTPException(404, "الجلسة غير موجودة")
    # تأكيد تلقائي للمطابقات عالية الثقة (>= 0.85) عند الإرسال — حتى لو لم يضغط المعلم تأكيد يدوياً
    participants = sess.get("participants", [])
    auto_confirmed_count = 0
    for p in participants:
        if (not p.get("confirmed")
                and not p.get("ignored")
                and p.get("matched_student_id")
                and (p.get("match_confidence") or 0) >= 0.85):
            p["confirmed"] = True
            auto_confirmed_count += 1
    update = {"status": "released", "released_at": now_iso()}
    if auto_confirmed_count:
        update["participants"] = participants
    await db.grade_sessions.update_one({"id": sid}, {"$set": update})
    ready = sum(1 for p in participants if p.get("confirmed") and not p.get("ignored") and p.get("matched_student_id"))
    return {"message": "تم إرسال الدرجات", "status": "released",
            "released_to": ready, "auto_confirmed": auto_confirmed_count}


@api_router.post("/grade-sessions/{sid}/reopen")
async def reopen_session(sid: str, t=Depends(get_teacher)):
    sess = await db.grade_sessions.find_one({"id": sid, "owner_id": t["teacher_id"]})
    if not sess:
        raise HTTPException(404, "الجلسة غير موجودة")
    await db.grade_sessions.update_one({"id": sid}, {"$set": {"status": "open", "released_at": None}})
    return {"message": "تمت إعادة الفتح", "status": "open"}


# ----------- Public (Student-facing) endpoints — لا تحتاج توثيق -----------

@api_router.get("/public/grade-sessions/{code}")
async def public_session_info(code: str):
    sess = await db.grade_sessions.find_one({"code": code.upper()}, {"_id": 0})
    if not sess:
        raise HTTPException(404, "الرابط غير صحيح أو منتهي")
    return _session_public_view(sess)


@api_router.post("/public/grade-sessions/{code}/join")
async def public_session_join(code: str, data: SessionJoinRequest):
    sess = await db.grade_sessions.find_one({"code": code.upper()})
    if not sess:
        raise HTTPException(404, "الرابط غير صحيح أو منتهي")
    name = data.name.strip()
    if not name:
        raise HTTPException(400, "اكتب اسمك")
    # تجنب التكرار: لو نفس الاسم سُجّل سابقاً يرجع نفس المشارك
    for p in sess.get("participants", []):
        if p["joined_name"].strip().lower() == name.lower():
            return {"participant_id": p["id"], "rejoined": True}
    gb = await db.gradebooks.find_one({"id": sess["gradebook_id"]}, {"_id": 0, "students": 1})
    match = _auto_match_participant(name, (gb or {}).get("students", []))
    # تأكيد تلقائي للمطابقات عالية الثقة (>= 0.85) — يلغي الحاجة لتأكيد يدوي للأسماء المطابقة تماماً
    auto_confirm = bool(match["matched_student_id"]) and (match["match_confidence"] or 0) >= 0.85
    participant = {
        "id": str(uuid.uuid4()),
        "joined_name": name,
        "joined_grade": data.grade.strip(),
        "joined_section": str(data.section).strip(),
        "matched_student_id": match["matched_student_id"],
        "matched_student_name": match["matched_student_name"],
        "match_confidence": match["match_confidence"],
        "confirmed": auto_confirm,
        "ignored": False,
        "joined_at": now_iso(),
    }
    await db.grade_sessions.update_one(
        {"id": sess["id"]},
        {"$push": {"participants": participant}}
    )
    return {"participant_id": participant["id"], "rejoined": False}


@api_router.get("/public/grade-sessions/{code}/state")
async def public_session_state(code: str, participant_id: str):
    sess = await db.grade_sessions.find_one({"code": code.upper()}, {"_id": 0})
    if not sess:
        raise HTTPException(404, "الرابط غير صحيح أو منتهي")
    p = next((x for x in sess.get("participants", []) if x["id"] == participant_id), None)
    if not p:
        raise HTTPException(404, "غير مسجل في الجلسة — انضم مجدداً")
    base = {
        "joined_name": p["joined_name"],
        "rubric_title": sess["rubric_title"],
        "column_label": sess["column_label"],
        "grade": sess["grade"],
        "section": sess["section"],
        "status": sess["status"],
    }
    if sess["status"] != "released":
        return {**base, "phase": "waiting"}
    if p.get("ignored") or not p.get("matched_student_id"):
        return {**base, "phase": "not_matched",
                "message": "لم يتم العثور على اسمك في السجل — راجع المعلم"}
    # جلب التقييم والمعايير
    rubric = await db.rubrics.find_one({"id": sess["rubric_id"]}, {"_id": 0})
    ev = await db.rubric_evaluations.find_one(
        {"rubric_id": sess["rubric_id"], "gradebook_id": sess["gradebook_id"], "student_id": p["matched_student_id"]},
        {"_id": 0}
    )
    if not rubric:
        return {**base, "phase": "not_matched", "message": "بطاقة التقييم محذوفة"}
    if not ev:
        return {**base, "phase": "not_evaluated",
                "message": "لم يتم تقييمك بعد على هذا النشاط"}
    ev_scores = ev.get("scores", {})
    criteria = [
        {"name": c["name"], "max": c["max"], "score": ev_scores.get(c["id"])}
        for c in rubric.get("criteria", [])
    ]
    gb = await db.gradebooks.find_one({"id": sess["gradebook_id"]}, {"_id": 0})
    return {
        **base,
        "phase": "released",
        "matched_name": p.get("matched_student_name"),
        "criteria": criteria,
        "total": ev.get("total"),
        "total_max": rubric.get("total_max"),
        "gb_score": ev.get("gb_score"),
        "gb_max": gb_max_map(gb).get(sess["column"]) if gb else None,
        "semester": sess["semester"],
    }


app.include_router(api_router)
app.add_middleware(
    CORSMiddleware, allow_credentials=True,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"]
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
